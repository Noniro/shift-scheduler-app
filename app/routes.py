from flask import Blueprint, render_template, request, redirect, url_for, jsonify, flash, current_app, session, make_response
from . import db
from .models import SchedulingPeriod, JobRole, ShiftDefinition, Worker, Constraint, ScheduledShift, User
from datetime import datetime, time, timedelta, date
from dateutil.parser import parse as parse_datetime
from sqlalchemy.orm import joinedload, selectinload # For eager loading

# For CSV export
import csv
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Option 1 - copilot 
"""
This file contains the main application routes for handling user interactions,
scheduling periods, job roles, workers, and shift assignments.

It includes routes for:
- Setting and clearing user names
- Managing scheduling periods
- Creating and managing job roles
- Generating shift slots
- Managing workers and their constraints
- Assigning shifts to workers based on constraints and qualifications
- etc...

This file is part of a Flask application that manages scheduling for workers in various job roles.

It uses SQLAlchemy for ORM and Flask's session management for user state.
It is designed to be modular, with routes grouped logically for easier maintenance and readability.
It also includes error handling and logging for better debugging and user feedback.
"""



# Option 2 - aistudion - google 
"""
routes.py - Defines the web routes and view functions for the Flask application.

This module is responsible for handling incoming HTTP requests and routing them
to the appropriate Python functions (view functions). These functions then
interact with the database models (from `app.models`), process user input,
manage session data, and render HTML templates to be sent back to the client's
browser.

The routes are organized using a Flask Blueprint named 'main_bp'.

Key functionalities handled by this module include:

1.  **User Identification:**
    -   Displaying the main page (`/`).
    -   Allowing users to set and clear their name, which is stored in the session.

2.  **Scheduling Period Management (`/periods`, etc.):**
    -   Creating new scheduling periods with start and end datetimes.
    -   Listing existing periods.
    -   Setting a period as "active" (stored in session), which contexts other operations.
    -   Deleting scheduling periods and their associated data (job roles, slots, assignments).

3.  **Job Role Management (within an active period) (`/period/<id>/roles`, etc.):**
    -   Adding job roles (e.g., "Cook", "Server") to the active scheduling period.
    -   Defining properties for each role like number needed and default shift duration.
    -   Listing and deleting job roles for the active period.

4.  **Coverage Slot Generation (`/period/<id>/generate_slots`):**
    -   Automatically generating individual `ShiftDefinition` (coverage slot) instances
        based on the job roles defined for the active period, their durations,
        and the period's start/end times.
    -   This process first clears any existing slots for the period.

5.  **Worker Management (`/manage_workers`, etc.):**
    -   Adding new workers to the system.
    -   Listing existing workers.
    -   Deleting workers.
    -   Editing a worker's qualified job roles for the currently active period.
    -   Adding unavailability constraints for workers (e.g., specific days they cannot work).

6.  **Shift Assignment (`/period/<id>/generate_slots_and_assign`):**
    -   A comprehensive action that combines:
        - Clearing old slots and assignments for the active period.
        - Generating new `ShiftDefinition` (coverage) slots.
        - Creating placeholder `ScheduledShift` objects for these new slots.
        - Invoking the shift assignment algorithm (`assign_shifts_fairly` from `app.algorithm`)
          to attempt to assign workers to the unassigned `ScheduledShift` placeholders.
    -   The main page (`/`) also displays the current schedule for the active period.

Helper Functions:
-   `get_active_period()`: Retrieves the currently active `SchedulingPeriod` object
    based on an ID stored in the user's session.

Common patterns used:
-   Flask's `render_template()` to display HTML pages.
-   `request.form` to access data from submitted HTML forms.
-   `redirect(url_for(...))` for Post/Redirect/Get pattern.
-   `flash()` messages for user feedback.
-   `session` object for storing user-specific data (user name, active period ID).
-   SQLAlchemy ORM for database interactions (queries, adding, committing, deleting objects).
-   Eager loading (`selectinload`, `joinedload`) for optimizing database queries,
    especially when displaying lists of related objects.
-   Error handling using try-except blocks and logging with `current_app.logger`.
"""

main_bp = Blueprint('main', __name__)

def get_active_period():
    active_period_id = session.get('active_period_id')
    if active_period_id:
        # FIX: Remove .options(selectinload(SchedulingPeriod.job_roles)) when using .get()
        # The lazy='dynamic' on job_roles means it's a query object already.
        # Accessing period.job_roles.count() later will execute an efficient count query.
        return SchedulingPeriod.query.get(active_period_id)
    return None


#     # Adding time restriction for job roles
# def is_time_within_role_restrictions(check_time, role):
#     """Check if a given time falls within the role's working hours"""
#     if not role.has_time_restrictions():
#         return True  # No restrictions means all times are valid
    
#     check_time_only = check_time.time()
    
#     if role.is_overnight_shift:
#         # For overnight shifts (e.g., 22:00 - 06:00)
#         # Valid if time >= start_time OR time <= end_time
#         return check_time_only >= role.work_start_time or check_time_only <= role.work_end_time
#     else:
#         # For same-day shifts (e.g., 09:00 - 17:00)
#         # Valid if start_time <= time <= end_time
#         return role.work_start_time <= check_time_only <= role.work_end_time


def is_time_within_role_restrictions(check_time, role):
    """Check if a given time falls within the role's working hours"""
    if not role.has_time_restrictions():
        return True  # No restrictions means all times are valid
    
    check_time_only = check_time.time()
    
    if role.is_overnight_shift:
        # For overnight shifts (e.g., 22:00 - 06:00)
        # Valid if time >= start_time OR time < end_time (not <=)
        return check_time_only >= role.work_start_time or check_time_only < role.work_end_time
    else:
        # For same-day shifts (e.g., 09:00 - 17:00)
        # Valid if start_time <= time < end_time (not <= for end)
        return role.work_start_time <= check_time_only < role.work_end_time


# --- User Name Routes ---
@main_bp.route('/', methods=['GET', 'POST'])
def index():
    current_user_name = session.get('user_name')
    if request.method == 'POST' and 'user_name_field' in request.form and not current_user_name:
        name_input = request.form.get('user_name_field')
        if name_input and name_input.strip(): # Ensure name is not just whitespace
            session['user_name'] = name_input.strip()
            session.permanent = True
            flash(f"Great to meet you, {session['user_name']}!", "success")
        else:
            flash("Please enter your name.", "warning")
        return redirect(url_for('main.index'))

    active_period = get_active_period()
    # Eager load qualified_roles for each worker
    workers = Worker.query.options(selectinload(Worker.qualified_roles)).order_by(Worker.name).all()
    
    scheduled_assignments = []
    has_defined_shift_slots = False # Slots are ShiftDefinition instances
    worker_hours = {} # NEW: Initialize worker hours dictionary

    if active_period:
        has_defined_shift_slots = ShiftDefinition.query.filter_by(scheduling_period_id=active_period.id).first() is not None
        if has_defined_shift_slots:
            scheduled_assignments = ScheduledShift.query.options(
                    joinedload(ScheduledShift.defined_slot).joinedload(ShiftDefinition.job_role), # Eager load job_role
                    joinedload(ScheduledShift.worker_assigned)
                ).join(ShiftDefinition).\
                filter(ShiftDefinition.scheduling_period_id == active_period.id).\
                order_by(ShiftDefinition.slot_start_datetime, ShiftDefinition.job_role_id, ShiftDefinition.instance_number).all()

            # NEW: Calculate worker hours from assignments
            if scheduled_assignments:
                for assignment in scheduled_assignments:
                    if assignment.worker_assigned:
                        worker = assignment.worker_assigned
                        # Calculate duration in hours
                        duration_hours = assignment.defined_slot.duration_timedelta.total_seconds() / 3600.0
                        worker_hours[worker.name] = worker_hours.get(worker.name, 0) + duration_hours


    return render_template('index.html',
                           current_user_name=current_user_name,
                           active_period=active_period,
                           workers=workers,
                           scheduled_assignments=scheduled_assignments,
                           has_defined_shift_slots=has_defined_shift_slots,
                           worker_hours=worker_hours) # NEW: Pass data to template

@main_bp.route('/set_user_name', methods=['POST'])
def set_user_name():
    user_name = request.form.get('user_name_field')
    if user_name and user_name.strip():
        session['user_name'] = user_name.strip(); session.permanent = True
        flash(f"Welcome, {user_name.strip()}! Your name is saved.", "success")
    else:
        flash("Please provide a name.", "warning")
    return redirect(url_for('main.index'))

@main_bp.route('/clear_name')
def clear_name():
    session.pop('user_name', None); flash("Your name has been cleared.", "info")
    return redirect(url_for('main.index'))

# --- Scheduling Period Routes ---
@main_bp.route('/periods', methods=['GET', 'POST'])
def manage_periods():
    if request.method == 'POST':
        try:
            name = request.form.get('period_name')
            start_date_str = request.form.get('period_start_date_hidden')
            end_date_str = request.form.get('period_end_date_hidden')
            start_time_str = request.form.get('period_start_time')
            end_time_str = request.form.get('period_end_time')

            if not name or not name.strip():
                flash("Period name is required.", "danger")
                return redirect(url_for('main.manage_periods'))
            name = name.strip()


            if not all([start_date_str, end_date_str, start_time_str, end_time_str]):
                flash("All period date and time fields are required.", "danger")
                return redirect(url_for('main.manage_periods'))

            if SchedulingPeriod.query.filter(SchedulingPeriod.name.ilike(name)).first():
                flash(f"A scheduling period with the name '{name}' already exists.", "danger")
                return redirect(url_for('main.manage_periods'))

            start_date_obj = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date_obj = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            start_time_obj = datetime.strptime(start_time_str, '%H:%M').time()
            end_time_obj = datetime.strptime(end_time_str, '%H:%M').time()

            if end_date_obj < start_date_obj or \
               (end_date_obj == start_date_obj and end_time_obj <= start_time_obj):
                flash("Period end must be after period start.", "danger")
                return redirect(url_for('main.manage_periods'))

            period_start_dt = datetime.combine(start_date_obj, start_time_obj)
            period_end_dt = datetime.combine(end_date_obj, end_time_obj)
            
            new_period = SchedulingPeriod(
                name=name, 
                period_start_datetime=period_start_dt, 
                period_end_datetime=period_end_dt
            )

            db.session.add(new_period); db.session.commit()
            flash(f"Scheduling Period '{name}' created. Now define its job roles.", "success")
            session['active_period_id'] = new_period.id
            session.permanent = True
            return redirect(url_for('main.manage_job_roles_for_period', period_id=new_period.id))
        except ValueError as ve:
            flash(f"Invalid date or time format: {ve}", "danger")
        except Exception as e:
            db.session.rollback(); flash(f"Error creating period: {e}", "danger")
            current_app.logger.error(f"Error creating period: {e}\n{request.form}")
        return redirect(url_for('main.manage_periods'))

    periods = SchedulingPeriod.query.order_by(SchedulingPeriod.name).all()
    active_period_id = session.get('active_period_id')
    return render_template('manage_periods.html', periods=periods, active_period_id=active_period_id)

@main_bp.route('/set_active_period/<int:period_id>', methods=['POST'])
def set_active_period_action(period_id):
    period_to_activate = SchedulingPeriod.query.get_or_404(period_id)
    session['active_period_id'] = period_id
    session.permanent = True
    flash(f"Period '{period_to_activate.name}' is now active. You can now define its job roles and shift slots.", "info")
    return redirect(url_for('main.manage_job_roles_for_period', period_id=period_id))

@main_bp.route('/delete_period/<int:period_id>', methods=['POST'])
def delete_period(period_id):
    period_to_delete = SchedulingPeriod.query.get_or_404(period_id)
    if session.get('active_period_id') == period_id:
        session.pop('active_period_id', None)
    db.session.delete(period_to_delete); db.session.commit()
    flash(f"Period '{period_to_delete.name}' and all its associated data deleted.", "success")
    return redirect(url_for('main.manage_periods'))

# # --- Job Role and Slot Generation Routes ---
# @main_bp.route('/period/<int:period_id>/roles', methods=['GET', 'POST'])
# def manage_job_roles_for_period(period_id):
#     period = SchedulingPeriod.query.get_or_404(period_id)
#     if session.get('active_period_id') != period_id:
#         session['active_period_id'] = period_id; session.permanent = True
#         flash(f"Active period set to '{period.name}'.", "info")
    
#     if request.method == 'POST':
#         try:
#             role_name = request.form.get('role_name')
#             number_needed_str = request.form.get('number_needed', '1')
#             days_str = request.form.get('duration_days', '0')
#             hours_str = request.form.get('duration_hours', '0')
#             minutes_str = request.form.get('duration_minutes', '0')
            
#             # UPDATED: Handle the new 1-5 difficulty range
#             difficulty_multiplier_str = request.form.get('difficulty_multiplier', '1')

#             # Time constraint fields
#             has_time_restrictions = request.form.get('has_time_restrictions') == 'on'
#             work_start_time_str = request.form.get('work_start_time')
#             work_end_time_str = request.form.get('work_end_time')
#             is_overnight_shift = request.form.get('is_overnight_shift') == 'on'
            
#             if not role_name or not role_name.strip(): 
#                 flash("Job role name is required.", "danger")
#             else:
#                 role_name = role_name.strip()
#                 number_needed = int(number_needed_str)
#                 days = int(days_str)
#                 hours = int(hours_str) 
#                 minutes = int(minutes_str)

#                 # UPDATED: Convert multiplier to int and validate range 1-5
#                 difficulty_multiplier = int(difficulty_multiplier_str)
#                 if difficulty_multiplier < 1 or difficulty_multiplier > 5:
#                     flash("Difficulty level must be between 1 and 5.", "danger")
#                     return redirect(url_for('main.manage_job_roles_for_period', period_id=period.id))
                
#                 if number_needed < 1: 
#                     flash("Number needed must be at least 1.", "danger")
#                 else:
#                     total_duration_minutes = (days * 24 * 60) + (hours * 60) + minutes
#                     if total_duration_minutes < 20: 
#                         flash("Minimum shift duration for a role is 20 minutes.", "danger")
#                     elif days < 0 or hours < 0 or minutes < 0 or hours >= 24 or minutes >= 60: 
#                         flash("Invalid duration values (e.g., hours 0-23, minutes 0-59).", "danger")
#                     elif JobRole.query.filter_by(scheduling_period_id=period.id, name=role_name).first(): 
#                         flash(f"Job role '{role_name}' already exists for this period.", "warning")
#                     else:
#                         # Parse time constraints if provided
#                         work_start_time = None
#                         work_end_time = None
                        
#                         if has_time_restrictions:
#                             if not work_start_time_str or not work_end_time_str:
#                                 flash("Both start and end times are required when restricting working hours.", "danger")
#                                 return redirect(url_for('main.manage_job_roles_for_period', period_id=period.id))
                            
#                             try:
#                                 work_start_time = datetime.strptime(work_start_time_str, '%H:%M').time()
#                                 work_end_time = datetime.strptime(work_end_time_str, '%H:%M').time()
                                
#                                 # Validate time logic
#                                 if not is_overnight_shift and work_end_time <= work_start_time:
#                                     flash("End time must be after start time for same-day shifts.", "danger")
#                                     return redirect(url_for('main.manage_job_roles_for_period', period_id=period.id))
#                                 elif is_overnight_shift and work_end_time >= work_start_time:
#                                     flash("For overnight shifts, end time should be earlier than start time (next day).", "warning")
                                    
#                             except ValueError:
#                                 flash("Invalid time format. Please use HH:MM format.", "danger")
#                                 return redirect(url_for('main.manage_job_roles_for_period', period_id=period.id))
                        
#                         new_role = JobRole(
#                             name=role_name, 
#                             number_needed=number_needed, 
#                             shift_duration_days=days, 
#                             shift_duration_hours=hours, 
#                             shift_duration_minutes=minutes,
#                             difficulty_multiplier=float(difficulty_multiplier), # Convert to float for database
#                             scheduling_period_id=period.id,
#                             work_start_time=work_start_time,
#                             work_end_time=work_end_time,
#                             is_overnight_shift=is_overnight_shift
#                         )
#                         db.session.add(new_role)
#                         db.session.commit()
                        
#                         time_info = ""
#                         if has_time_restrictions:
#                             time_info = f" (Working hours: {work_start_time_str} - {work_end_time_str}{'next day' if is_overnight_shift else ''})"
                        
#                         difficulty_labels = {1: "Easy/Regular", 2: "Light", 3: "Moderate", 4: "Hard", 5: "Very Hard"}
#                         flash(f"Job Role '{role_name}' added with difficulty level {difficulty_multiplier} ({difficulty_labels[difficulty_multiplier]}).{time_info}", "success")
                        
#         except ValueError: 
#             flash("Invalid number for 'Needed', 'Duration' or 'Difficulty' fields.", "danger")
#         except Exception as e: 
#             db.session.rollback()
#             flash(f"Error adding job role: {e}", "danger")
#             current_app.logger.error(f"Error adding job role for period {period.id}: {e}\n{request.form}")
#         return redirect(url_for('main.manage_job_roles_for_period', period_id=period.id))


#     job_roles = JobRole.query.filter_by(scheduling_period_id=period.id).order_by(JobRole.name).all()
#     generated_slots = ShiftDefinition.query.options(joinedload(ShiftDefinition.job_role))\
#                                            .filter_by(scheduling_period_id=period.id)\
#                                            .order_by(ShiftDefinition.job_role_id, ShiftDefinition.instance_number, ShiftDefinition.slot_start_datetime).all()
#     has_generated_slots = bool(generated_slots)
#     workers_exist = Worker.query.first() is not None
#     can_assign = workers_exist and job_roles

#     # ---- RETRIEVE DETAILED MESSAGES FROM SESSION ----
#     assignment_details = session.pop('assignment_details', None) # Get and remove from session
#     # ---- END OF RETRIEVAL ----

#     return render_template('manage_job_roles.html', 
#                            period=period, 
#                            job_roles=job_roles, 
#                            generated_slots=generated_slots,
#                            has_generated_slots=has_generated_slots,
#                            can_assign=can_assign,
#                            workers_exist=workers_exist,
#                            assignment_details=assignment_details) # Pass to template



@main_bp.route('/period/<int:period_id>/roles', methods=['GET', 'POST'])
def manage_job_roles_for_period(period_id):
    period = SchedulingPeriod.query.get_or_404(period_id)
    if session.get('active_period_id') != period_id:
        session['active_period_id'] = period_id; session.permanent = True
        flash(f"Active period set to '{period.name}'.", "info")
    
    if request.method == 'POST':
        try:
            role_name = request.form.get('role_name')
            number_needed_str = request.form.get('number_needed', '1')
            days_str = request.form.get('duration_days', '0')
            hours_str = request.form.get('duration_hours', '0')
            minutes_str = request.form.get('duration_minutes', '0')
            
            # REMOVED: Difficulty multiplier processing

            # Time constraint fields
            has_time_restrictions = request.form.get('has_time_restrictions') == 'on'
            work_start_time_str = request.form.get('work_start_time')
            work_end_time_str = request.form.get('work_end_time')
            is_overnight_shift = request.form.get('is_overnight_shift') == 'on'
            
            if not role_name or not role_name.strip(): 
                flash("Job role name is required.", "danger")
            else:
                role_name = role_name.strip()
                number_needed = int(number_needed_str)
                days = int(days_str)
                hours = int(hours_str) 
                minutes = int(minutes_str)

                # REMOVED: Difficulty multiplier validation
                
                if number_needed < 1: 
                    flash("Number needed must be at least 1.", "danger")
                else:
                    total_duration_minutes = (days * 24 * 60) + (hours * 60) + minutes
                    if total_duration_minutes < 20: 
                        flash("Minimum shift duration for a role is 20 minutes.", "danger")
                    elif days < 0 or hours < 0 or minutes < 0 or hours >= 24 or minutes >= 60: 
                        flash("Invalid duration values (e.g., hours 0-23, minutes 0-59).", "danger")
                    elif JobRole.query.filter_by(scheduling_period_id=period.id, name=role_name).first(): 
                        flash(f"Job role '{role_name}' already exists for this period.", "warning")
                    else:
                        # Parse time constraints if provided
                        work_start_time = None
                        work_end_time = None
                        
                        if has_time_restrictions:
                            if not work_start_time_str or not work_end_time_str:
                                flash("Both start and end times are required when restricting working hours.", "danger")
                                return redirect(url_for('main.manage_job_roles_for_period', period_id=period.id))
                            
                            try:
                                work_start_time = datetime.strptime(work_start_time_str, '%H:%M').time()
                                work_end_time = datetime.strptime(work_end_time_str, '%H:%M').time()
                                
                                # Validate time logic
                                if not is_overnight_shift and work_end_time <= work_start_time:
                                    flash("End time must be after start time for same-day shifts.", "danger")
                                    return redirect(url_for('main.manage_job_roles_for_period', period_id=period.id))
                                elif is_overnight_shift and work_end_time >= work_start_time:
                                    flash("For overnight shifts, end time should be earlier than start time (next day).", "warning")
                                    
                            except ValueError:
                                flash("Invalid time format. Please use HH:MM format.", "danger")
                                return redirect(url_for('main.manage_job_roles_for_period', period_id=period.id))
                        
                        new_role = JobRole(
                            name=role_name, 
                            number_needed=number_needed, 
                            shift_duration_days=days, 
                            shift_duration_hours=hours, 
                            shift_duration_minutes=minutes,
                            difficulty_multiplier=1.0, # SET DEFAULT VALUE
                            scheduling_period_id=period.id,
                            work_start_time=work_start_time,
                            work_end_time=work_end_time,
                            is_overnight_shift=is_overnight_shift
                        )
                        db.session.add(new_role)
                        db.session.commit()
                        
                        time_info = ""
                        if has_time_restrictions:
                            time_info = f" (Working hours: {work_start_time_str} - {work_end_time_str}{'next day' if is_overnight_shift else ''})"
                        
                        # UPDATED: Removed difficulty reference from success message
                        flash(f"Job Role '{role_name}' added successfully.{time_info}", "success")
                        
        except ValueError: 
            # UPDATED: Removed 'Difficulty' from error message
            flash("Invalid number for 'Needed' or 'Duration' fields.", "danger")
        except Exception as e: 
            db.session.rollback()
            flash(f"Error adding job role: {e}", "danger")
            current_app.logger.error(f"Error adding job role for period {period.id}: {e}\n{request.form}")
        return redirect(url_for('main.manage_job_roles_for_period', period_id=period.id))


    job_roles = JobRole.query.filter_by(scheduling_period_id=period.id).order_by(JobRole.name).all()
    generated_slots = ShiftDefinition.query.options(joinedload(ShiftDefinition.job_role))\
                                           .filter_by(scheduling_period_id=period.id)\
                                           .order_by(ShiftDefinition.job_role_id, ShiftDefinition.instance_number, ShiftDefinition.slot_start_datetime).all()
    has_generated_slots = bool(generated_slots)
    workers_exist = Worker.query.first() is not None
    can_assign = workers_exist and job_roles

    # ---- RETRIEVE DETAILED MESSAGES FROM SESSION ----
    assignment_details = session.pop('assignment_details', None) # Get and remove from session
    # ---- END OF RETRIEVAL ----

    return render_template('manage_job_roles.html', 
                           period=period, 
                           job_roles=job_roles, 
                           generated_slots=generated_slots,
                           has_generated_slots=has_generated_slots,
                           can_assign=can_assign,
                           workers_exist=workers_exist,
                           assignment_details=assignment_details) # Pass to template





@main_bp.route('/period/<int:period_id>/role/<int:role_id>/delete', methods=['POST'])
def delete_job_role(period_id, role_id):
    role = JobRole.query.filter_by(id=role_id, scheduling_period_id=period_id).first_or_404()
    db.session.delete(role); db.session.commit()
    flash(f"Job Role '{role.name}' and its generated slots/assignments deleted.", "info")
    return redirect(url_for('main.manage_job_roles_for_period', period_id=period_id))



@main_bp.route('/period/<int:period_id>/role/<int:role_id>/edit', methods=['GET', 'POST'])
def edit_job_role(period_id, role_id):
    """Edit an existing job role for a scheduling period"""
    period = SchedulingPeriod.query.get_or_404(period_id)
    role = JobRole.query.filter_by(id=role_id, scheduling_period_id=period_id).first_or_404()

    if request.method == 'POST':
        try:
            role_name = request.form.get('role_name')
            number_needed_str = request.form.get('number_needed', '1')
            days_str = request.form.get('duration_days', '0')
            hours_str = request.form.get('duration_hours', '0')
            minutes_str = request.form.get('duration_minutes', '0')
            
            # REMOVED: Difficulty multiplier processing

            has_time_restrictions = request.form.get('has_time_restrictions') == 'on'
            work_start_time_str = request.form.get('work_start_time')
            work_end_time_str = request.form.get('work_end_time')
            is_overnight_shift = request.form.get('is_overnight_shift') == 'on'

            if not role_name or not role_name.strip():
                flash("Job role name is required.", "danger")
                return redirect(url_for('main.edit_job_role', period_id=period_id, role_id=role_id))

            role_name = role_name.strip()
            number_needed = int(number_needed_str)
            days = int(days_str)
            hours = int(hours_str)
            minutes = int(minutes_str)
            
            # REMOVED: Difficulty multiplier validation

            if number_needed < 1:
                flash("Number needed must be at least 1.", "danger")
                return redirect(url_for('main.edit_job_role', period_id=period_id, role_id=role_id))

            total_duration_minutes = (days * 24 * 60) + (hours * 60) + minutes
            if total_duration_minutes < 20:
                flash("Minimum shift duration for a role is 20 minutes.", "danger")
                return redirect(url_for('main.edit_job_role', period_id=period_id, role_id=role_id))
            if days < 0 or hours < 0 or minutes < 0 or hours >= 24 or minutes >= 60:
                flash("Invalid duration values (e.g., hours 0-23, minutes 0-59).", "danger")
                return redirect(url_for('main.edit_job_role', period_id=period_id, role_id=role_id))

            # Check if another role with same name exists in this period
            existing_role = JobRole.query.filter(
                JobRole.scheduling_period_id == period_id,
                JobRole.name.ilike(role_name),
                JobRole.id != role_id
            ).first()
            if existing_role:
                flash(f"Job role '{role_name}' already exists for this period.", "danger")
                return redirect(url_for('main.edit_job_role', period_id=period_id, role_id=role_id))

            work_start_time = None
            work_end_time = None
            if has_time_restrictions:
                if not work_start_time_str or not work_end_time_str:
                    flash("Both start and end times are required when restricting working hours.", "danger")
                    return redirect(url_for('main.edit_job_role', period_id=period_id, role_id=role_id))
                try:
                    work_start_time = datetime.strptime(work_start_time_str, '%H:%M').time()
                    work_end_time = datetime.strptime(work_end_time_str, '%H:%M').time()
                    if not is_overnight_shift and work_end_time <= work_start_time:
                        flash("End time must be after start time for same-day shifts.", "danger")
                        return redirect(url_for('main.edit_job_role', period_id=period_id, role_id=role_id))
                    elif is_overnight_shift and work_end_time >= work_start_time:
                        flash("For overnight shifts, end time should be earlier than start time (next day).", "warning")
                except ValueError:
                    flash("Invalid time format. Please use HH:MM format.", "danger")
                    return redirect(url_for('main.edit_job_role', period_id=period_id, role_id=role_id))

            # Update role
            role.name = role_name
            role.number_needed = number_needed
            role.shift_duration_days = days
            role.shift_duration_hours = hours
            role.shift_duration_minutes = minutes
            # REMOVED: difficulty_multiplier update (keep existing value)
            role.work_start_time = work_start_time
            role.work_end_time = work_end_time
            role.is_overnight_shift = is_overnight_shift if has_time_restrictions else False

            db.session.commit()
            # UPDATED: Removed difficulty reference from success message
            flash(f"Job Role '{role.name}' updated successfully.", "success")
            return redirect(url_for('main.manage_job_roles_for_period', period_id=period_id))

        except Exception as e:
            db.session.rollback()
            flash(f"Error updating job role: {e}", "danger")
            current_app.logger.error(f"Error updating job role {role_id} for period {period_id}: {e}\n{request.form}")
            return redirect(url_for('main.edit_job_role', period_id=period_id, role_id=role_id))

    # GET request
    return render_template('edit_job_role.html', period=period, role=role)

# --- Job Role Editing Route ---
# This route allows editing an existing job role for a scheduling period.

# @main_bp.route('/period/<int:period_id>/role/<int:role_id>/edit', methods=['GET', 'POST'])
# def edit_job_role(period_id, role_id):
#     """Edit an existing job role for a scheduling period"""
#     period = SchedulingPeriod.query.get_or_404(period_id)
#     role = JobRole.query.filter_by(id=role_id, scheduling_period_id=period_id).first_or_404()

#     if request.method == 'POST':
#         try:
#             role_name = request.form.get('role_name')
#             number_needed_str = request.form.get('number_needed', '1')
#             days_str = request.form.get('duration_days', '0')
#             hours_str = request.form.get('duration_hours', '0')
#             minutes_str = request.form.get('duration_minutes', '0')
            
#             # UPDATED: Handle the new 1-5 difficulty range
#             difficulty_multiplier_str = request.form.get('difficulty_multiplier', '1')

#             has_time_restrictions = request.form.get('has_time_restrictions') == 'on'
#             work_start_time_str = request.form.get('work_start_time')
#             work_end_time_str = request.form.get('work_end_time')
#             is_overnight_shift = request.form.get('is_overnight_shift') == 'on'

#             if not role_name or not role_name.strip():
#                 flash("Job role name is required.", "danger")
#                 return redirect(url_for('main.edit_job_role', period_id=period_id, role_id=role_id))

#             role_name = role_name.strip()
#             number_needed = int(number_needed_str)
#             days = int(days_str)
#             hours = int(hours_str)
#             minutes = int(minutes_str)
            
#             # UPDATED: Convert multiplier to int and validate range 1-5
#             difficulty_multiplier = int(difficulty_multiplier_str)
#             if difficulty_multiplier < 1 or difficulty_multiplier > 5:
#                 flash("Difficulty level must be between 1 and 5.", "danger")
#                 return redirect(url_for('main.edit_job_role', period_id=period_id, role_id=role_id))

#             if number_needed < 1:
#                 flash("Number needed must be at least 1.", "danger")
#                 return redirect(url_for('main.edit_job_role', period_id=period_id, role_id=role_id))

#             total_duration_minutes = (days * 24 * 60) + (hours * 60) + minutes
#             if total_duration_minutes < 20:
#                 flash("Minimum shift duration for a role is 20 minutes.", "danger")
#                 return redirect(url_for('main.edit_job_role', period_id=period_id, role_id=role_id))
#             if days < 0 or hours < 0 or minutes < 0 or hours >= 24 or minutes >= 60:
#                 flash("Invalid duration values (e.g., hours 0-23, minutes 0-59).", "danger")
#                 return redirect(url_for('main.edit_job_role', period_id=period_id, role_id=role_id))

#             # Check if another role with same name exists in this period
#             existing_role = JobRole.query.filter(
#                 JobRole.scheduling_period_id == period_id,
#                 JobRole.name.ilike(role_name),
#                 JobRole.id != role_id
#             ).first()
#             if existing_role:
#                 flash(f"Job role '{role_name}' already exists for this period.", "danger")
#                 return redirect(url_for('main.edit_job_role', period_id=period_id, role_id=role_id))

#             work_start_time = None
#             work_end_time = None
#             if has_time_restrictions:
#                 if not work_start_time_str or not work_end_time_str:
#                     flash("Both start and end times are required when restricting working hours.", "danger")
#                     return redirect(url_for('main.edit_job_role', period_id=period_id, role_id=role_id))
#                 try:
#                     work_start_time = datetime.strptime(work_start_time_str, '%H:%M').time()
#                     work_end_time = datetime.strptime(work_end_time_str, '%H:%M').time()
#                     if not is_overnight_shift and work_end_time <= work_start_time:
#                         flash("End time must be after start time for same-day shifts.", "danger")
#                         return redirect(url_for('main.edit_job_role', period_id=period_id, role_id=role_id))
#                     elif is_overnight_shift and work_end_time >= work_start_time:
#                         flash("For overnight shifts, end time should be earlier than start time (next day).", "warning")
#                 except ValueError:
#                     flash("Invalid time format. Please use HH:MM format.", "danger")
#                     return redirect(url_for('main.edit_job_role', period_id=period_id, role_id=role_id))

#             # Update role
#             role.name = role_name
#             role.number_needed = number_needed
#             role.shift_duration_days = days
#             role.shift_duration_hours = hours
#             role.shift_duration_minutes = minutes
#             role.difficulty_multiplier = float(difficulty_multiplier)  # Convert to float for database
#             role.work_start_time = work_start_time
#             role.work_end_time = work_end_time
#             role.is_overnight_shift = is_overnight_shift if has_time_restrictions else False

#             db.session.commit()
            
#             difficulty_labels = {1: "Easy/Regular", 2: "Light", 3: "Moderate", 4: "Hard", 5: "Very Hard"}
#             flash(f"Job Role '{role.name}' updated with difficulty level {difficulty_multiplier} ({difficulty_labels[difficulty_multiplier]}).", "success")
#             return redirect(url_for('main.manage_job_roles_for_period', period_id=period_id))

#         except Exception as e:
#             db.session.rollback()
#             flash(f"Error updating job role: {e}", "danger")
#             current_app.logger.error(f"Error updating job role {role_id} for period {period_id}: {e}\n{request.form}")
#             return redirect(url_for('main.edit_job_role', period_id=period_id, role_id=role_id))

#     # GET request
#     return render_template('edit_job_role.html', period=period, role=role)


# --- Worker and Constraint Routes ---
@main_bp.route('/manage_workers', methods=['GET', 'POST'])
def manage_workers():
    active_period = get_active_period()
    all_job_roles_in_active_period = []
    if active_period:
        all_job_roles_in_active_period = JobRole.query.filter_by(scheduling_period_id=active_period.id).order_by(JobRole.name).all()
    
    if request.method == 'POST':
        # ... (POST logic remains the same as your last full version) ...
        name = request.form.get('worker_name')
        email_from_form = request.form.get('worker_email')
        max_hours_str = request.form.get('max_hours_per_week')
        max_hours = int(max_hours_str) if max_hours_str and max_hours_str.isdigit() else None
        qualified_role_ids_str = request.form.getlist('qualified_roles') 

        if not name or not name.strip():
            flash('Worker name is required.', 'danger'); return redirect(url_for('main.manage_workers'))
        name = name.strip()

        processed_email = email_from_form.strip() if email_from_form else None
        if not processed_email: processed_email = None

        if Worker.query.filter(Worker.name.ilike(name)).first():
            flash(f'Worker with name "{name}" already exists.', 'warning'); return redirect(url_for('main.manage_workers'))
        if processed_email and Worker.query.filter(Worker.email.ilike(processed_email)).first():
            flash(f'Worker with email "{processed_email}" already exists.', 'warning'); return redirect(url_for('main.manage_workers'))
        
        try:
            new_worker = Worker(name=name, email=processed_email, max_hours_per_week=max_hours)
            if active_period and qualified_role_ids_str:
                qualified_role_ids = [int(r_id) for r_id in qualified_role_ids_str]
                roles_to_assign = JobRole.query.filter(
                    JobRole.id.in_(qualified_role_ids),
                    JobRole.scheduling_period_id == active_period.id 
                ).all()
                for role in roles_to_assign:
                    new_worker.qualified_roles.append(role)
            
            db.session.add(new_worker); db.session.commit()
            flash(f'Worker "{name}" added successfully.', 'success')
        except Exception as e:
            db.session.rollback(); flash(f'Error adding worker: {e}', 'danger')
            current_app.logger.error(f"Error adding worker {name}: {e}")
        return redirect(url_for('main.manage_workers'))

    # ---- FIX IS HERE for the GET request part ----
    workers = Worker.query.options(
        selectinload(Worker.qualified_roles).selectinload(JobRole.scheduling_period)
        # REMOVE: selectinload(Worker.constraints) 
    ).order_by(Worker.name).all()
    # ---- END OF FIX ----
    
    return render_template('manage_workers.html', workers=workers, active_period=active_period, 
                           all_job_roles_in_active_period=all_job_roles_in_active_period)

@main_bp.route('/worker/<int:worker_id>/delete', methods=['POST'])
def delete_worker(worker_id):
    worker = Worker.query.get_or_404(worker_id)
    db.session.delete(worker); db.session.commit()
    flash(f"Worker '{worker.name}' deleted.", "info")
    return redirect(url_for('main.manage_workers'))


@main_bp.route('/worker/<int:worker_id>/edit_roles', methods=['POST'])
def edit_worker_roles(worker_id):
    worker = Worker.query.options(selectinload(Worker.qualified_roles)).get_or_404(worker_id)
    active_period = get_active_period()
    if not active_period:
        flash("No active period to manage roles for.", "warning"); return redirect(url_for('main.manage_workers'))

    submitted_role_ids = set(map(int, request.form.getlist('qualified_roles')))
    
    current_period_roles_for_worker = {role.id for role in worker.qualified_roles if role.scheduling_period_id == active_period.id}
    
    for role_id_to_remove in current_period_roles_for_worker - submitted_role_ids:
        role = JobRole.query.get(role_id_to_remove)
        if role and role in worker.qualified_roles:
            worker.qualified_roles.remove(role)
            
    for role_id_to_add in submitted_role_ids - current_period_roles_for_worker:
        role = JobRole.query.filter_by(id=role_id_to_add, scheduling_period_id=active_period.id).first()
        if role and role not in worker.qualified_roles:
             worker.qualified_roles.append(role)
             
    try:
        db.session.commit(); flash(f"Roles updated for worker {worker.name} for period '{active_period.name}'.", "success")
    except Exception as e:
        db.session.rollback(); flash(f"Error updating roles: {e}", "danger")
        current_app.logger.error(f"Error updating roles for worker {worker.id}: {e}")
    return redirect(url_for('main.manage_workers'))


# Replace your add_constraint route in routes.py with this updated version:

@main_bp.route('/worker/<int:worker_id>/add_constraint', methods=['POST'])
def add_constraint(worker_id):
    worker = Worker.query.get_or_404(worker_id)
    target_redirect = request.form.get('redirect_to', url_for('main.manage_workers'))
    
    try:
        constraint_type = request.form.get('constraint_type')  # 'full_day' or 'specific_hours'
        description = request.form.get('constraint_description', '').strip() or None
        
        if constraint_type == 'full_day':
            # Handle full day constraints (existing logic)
            start_date_str = request.form.get('constraint_start_date')
            end_date_str = request.form.get('constraint_end_date')
            
            if not start_date_str or not end_date_str:
                flash("Both start and end dates for unavailability are required.", "danger")
                return redirect(target_redirect)
                
            start_date_obj = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date_obj = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            
            if end_date_obj < start_date_obj:
                flash("End date cannot be before start date.", "danger")
                return redirect(target_redirect)
                
            # Set to full day (00:00 to 23:59)
            cs_dt = datetime.combine(start_date_obj, time.min)
            ce_dt = datetime.combine(end_date_obj, time.max)
            constraint_type_db = "UNAVAILABLE_DAY_RANGE"
            
        elif constraint_type == 'specific_hours':
            # Handle specific hours constraints (new functionality)
            start_date_str = request.form.get('start_datetime_date')
            start_time_str = request.form.get('start_datetime_time')
            end_date_str = request.form.get('end_datetime_date')
            end_time_str = request.form.get('end_datetime_time')
            
            if not all([start_date_str, start_time_str, end_date_str, end_time_str]):
                flash("All date and time fields are required for specific hours constraints.", "danger")
                return redirect(target_redirect)
                
            start_date_obj = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            start_time_obj = datetime.strptime(start_time_str, '%H:%M').time()
            end_date_obj = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            end_time_obj = datetime.strptime(end_time_str, '%H:%M').time()
            
            cs_dt = datetime.combine(start_date_obj, start_time_obj)
            ce_dt = datetime.combine(end_date_obj, end_time_obj)
            
            if ce_dt <= cs_dt:
                flash("End date/time must be after start date/time.", "danger")
                return redirect(target_redirect)
                
            constraint_type_db = "UNAVAILABLE_TIME_RANGE"
            
        else:
            flash("Invalid constraint type.", "danger")
            return redirect(target_redirect)
        
        # Create the constraint
        constraint = Constraint(
            worker_id=worker.id, 
            constraint_type=constraint_type_db, 
            start_datetime=cs_dt, 
            end_datetime=ce_dt,
            description=description
        )
        
        db.session.add(constraint)
        db.session.commit()
        
        # Create success message
        duration_info = constraint.get_duration_str()
        constraint_desc = constraint.get_constraint_description()
        flash(f'Constraint added for {worker.name}: {constraint_desc} ({duration_info})', 'success')
        
    except ValueError as ve:
        flash(f"Invalid date/time format: {ve}", "danger")
    except Exception as e:
        db.session.rollback()
        flash(f'Error adding constraint: {e}', 'danger')
        current_app.logger.error(f"Error in add_constraint for worker {worker_id}: {e}\n{request.form}")
    
    return redirect(target_redirect)

# Add this new route for deleting constraints:
@main_bp.route('/constraint/<int:constraint_id>/delete', methods=['POST'])
def delete_constraint(constraint_id):
    constraint = Constraint.query.get_or_404(constraint_id)
    worker_name = constraint.worker.name
    constraint_desc = constraint.get_constraint_description()
    
    db.session.delete(constraint)
    db.session.commit()
    
    flash(f'Constraint deleted for {worker_name}: {constraint_desc}', 'info')
    return redirect(url_for('main.manage_workers'))

# Assignments and Slot Generation Routes according to the algorithm











################ WITH DEBUGG MODE ###################

import random

# @main_bp.route('/period/<int:period_id>/generate_slots_and_assign', methods=['POST'])
# def generate_slots_and_assign_action(period_id):
#     period = SchedulingPeriod.query.get_or_404(period_id)
    
#     # ============ RANDOM SEED HANDLING ============
#     random_seed = request.form.get('random_seed', '').strip()
#     if random_seed:
#         try:
#             random_seed = int(random_seed)
#             if random_seed < 1 or random_seed > 999999:
#                 flash("Random seed must be between 1 and 999999. Using auto-generated seed.", "warning")
#                 random_seed = random.randint(1, 999999)
#         except ValueError:
#             flash("Invalid random seed. Using auto-generated seed.", "warning")
#             random_seed = random.randint(1, 999999)
#     else:
#         random_seed = random.randint(1, 999999)
    
#     # Set the seed for reproducible results
#     random.seed(random_seed)
#     current_app.logger.info(f"Using random seed: {random_seed}")
    
#     # === DEBUGGING CONFIGURATION ===
#     DEBUG_SLOT_GENERATION = True  # Set to False to disable all debugging
#     DEBUG_ROLE_NAMES = []  # Empty list = debug ALL roles, or specify: ["Toran", "Cook", "Guard"]
#     DEBUG_MAX_ITERATIONS_TO_SHOW = 10  # Only show first N iterations per role to avoid spam
    
#     current_app.logger.info(f"Clearing old data for period {period.id} ('{period.name}')")
#     ids_to_delete_assignments = [s.id for s in ScheduledShift.query.join(ShiftDefinition)
#                                .filter(ShiftDefinition.scheduling_period_id == period_id).all()]
#     if ids_to_delete_assignments:
#         ScheduledShift.query.filter(ScheduledShift.id.in_(ids_to_delete_assignments)).delete(synchronize_session=False)
#     ShiftDefinition.query.filter_by(scheduling_period_id=period.id).delete()
#     db.session.commit()
#     current_app.logger.info(f"Old data cleared for period {period.id}.")

#     job_roles_for_period = JobRole.query.filter_by(scheduling_period_id=period.id).all()
#     if not job_roles_for_period:
#         flash("No job roles defined for this period. Cannot generate slots or assign.", "warning")
#         return redirect(url_for('main.manage_job_roles_for_period', period_id=period.id))

#     # ============ RANDOMIZE JOB ROLE ORDER ============
#     random.shuffle(job_roles_for_period)
#     current_app.logger.info(f"Randomized processing order for {len(job_roles_for_period)} job roles")

#     total_new_slots_generated = 0
#     generated_slot_objects = []
    
#     for role in job_roles_for_period:
#         # === GENERIC DEBUG CHECK ===
#         should_debug_this_role = (
#             DEBUG_SLOT_GENERATION and 
#             (not DEBUG_ROLE_NAMES or role.name in DEBUG_ROLE_NAMES)
#         )
        
#         if should_debug_this_role:
#             print(f"\n{'='*60}")
#             print(f"DEBUGGING ROLE: {role.name}")
#             print(f"{'='*60}")
#             print(f"Configuration:")
#             print(f"  - Number needed: {role.number_needed}")
#             print(f"  - Duration: {role.get_duration_timedelta()}")
#             print(f"  - Has time restrictions: {role.has_time_restrictions()}")
#             if role.has_time_restrictions():
#                 print(f"  - Work hours: {role.work_start_time} - {role.work_end_time}")
#                 print(f"  - Is overnight: {role.is_overnight_shift}")
#             print(f"  - Period: {period.period_start_datetime} to {period.period_end_datetime}")
#             print(f"  - Difficulty multiplier: {role.difficulty_multiplier}")
#             print(f"  - Random seed: {random_seed}")
#             print("-" * 60)
        
#         role_slots_generated_this_role = 0
#         current_dt_for_role = period.period_start_datetime
#         duration = role.get_duration_timedelta()
        
#         if duration.total_seconds() <= 0:
#             current_app.logger.warning(f"Skipping role '{role.name}' due to zero duration for period {period.id}.")
#             flash(f"Job Role '{role.name}' has zero/negative shift duration and was skipped for slot generation.", "warning")
#             continue
            
#         # SAFETY: Prevent infinite loops, but allow legitimate high iteration counts
#         max_iter = 5000  
#         iter_count = 0
#         iterations_shown = 0
#         consecutive_invalid_slots = 0  # Track consecutive failures to detect infinite loops
        
#         while current_dt_for_role < period.period_end_datetime and iter_count < max_iter:
#             iter_count += 1
#             show_this_iteration = (
#                 should_debug_this_role and 
#                 iterations_shown < DEBUG_MAX_ITERATIONS_TO_SHOW
#             )
            
#             if show_this_iteration:
#                 print(f"\nIteration {iter_count}:")
#                 print(f"  Current time: {current_dt_for_role}")
            
#             # Check if current time is within role's working hours
#             if role.has_time_restrictions():
#                 is_valid_time = is_time_within_role_restrictions(current_dt_for_role, role)
                
#                 if show_this_iteration:
#                     print(f"  Time restriction check: {is_valid_time}")
#                     print(f"  Current time only: {current_dt_for_role.time()}")
#                     print(f"  Work window: {role.work_start_time} - {role.work_end_time}")
                
#                 if not is_valid_time:
#                     if show_this_iteration:
#                         print(f"  SKIPPING: Time not within restrictions")
                    
#                     # FIXED: Move to next valid time slot properly
#                     current_dt_for_role = get_next_valid_start_time(current_dt_for_role, role, period.period_end_datetime)
#                     consecutive_invalid_slots = 0  # Reset counter
                    
#                     if show_this_iteration:
#                         print(f"  Moved to next valid time: {current_dt_for_role}")
#                     continue
            
#             slot_start = current_dt_for_role
#             slot_end = current_dt_for_role + duration
            
#             if show_this_iteration:
#                 print(f"  Valid time - creating slot: {slot_start} to {slot_end}")
            
#             # For time-restricted roles, ensure slot doesn't exceed working hours
#             original_slot_end = slot_end
#             if role.has_time_restrictions():
#                 slot_end = constrain_slot_to_working_hours(slot_start, slot_end, role)
                
#                 if show_this_iteration and slot_end != original_slot_end:
#                     print(f"  Time constraint: Adjusted end from {original_slot_end} to {slot_end}")
            
#             # Ensure slot doesn't exceed period end
#             if slot_end > period.period_end_datetime:
#                 original_slot_end = slot_end
#                 slot_end = period.period_end_datetime
#                 if show_this_iteration:
#                     print(f"  Period limit: Adjusted end from {original_slot_end} to {slot_end}")
            
#             if slot_start < slot_end:
#                 # Valid slot - create it
#                 if show_this_iteration:
#                     print(f"  CREATING {role.number_needed} slots from {slot_start} to {slot_end}")
                
#                 for i in range(1, role.number_needed + 1):
#                     new_slot = ShiftDefinition(
#                         slot_start_datetime=slot_start, 
#                         slot_end_datetime=slot_end,
#                         instance_number=i, 
#                         scheduling_period_id=period.id, 
#                         job_role_id=role.id
#                     )
#                     db.session.add(new_slot)
#                     generated_slot_objects.append(new_slot)
#                     role_slots_generated_this_role += 1
                    
#                     if show_this_iteration:
#                         print(f"    Created slot #{i}: {new_slot.name}")
                        
#                 if show_this_iteration:
#                     iterations_shown += 1
                
#                 # FIXED: Always advance by the ORIGINAL duration, not the constrained slot_end
#                 current_dt_for_role = slot_start + duration
#                 consecutive_invalid_slots = 0  # Reset counter
                
#             else:
#                 # Invalid slot - need to advance time properly
#                 consecutive_invalid_slots += 1
                
#                 if show_this_iteration:
#                     print(f"  INVALID SLOT: start >= end ({slot_start} >= {slot_end})")
                
#                 # FIXED: Handle infinite loop prevention
#                 if consecutive_invalid_slots >= 10:
#                     if show_this_iteration:
#                         print(f"  BREAKING: Too many consecutive invalid slots, moving to next day")
                    
#                     # Force advance to next day at start of working hours
#                     if role.has_time_restrictions():
#                         next_day = current_dt_for_role.replace(
#                             hour=role.work_start_time.hour,
#                             minute=role.work_start_time.minute,
#                             second=0,
#                             microsecond=0
#                         ) + timedelta(days=1)
#                         current_dt_for_role = next_day
#                     else:
#                         # For unrestricted roles, advance by minimum meaningful time
#                         current_dt_for_role += timedelta(hours=1)
                    
#                     consecutive_invalid_slots = 0
#                     continue
                
#                 # Try smaller advancement
#                 if role.has_time_restrictions():
#                     current_dt_for_role = get_next_valid_start_time(current_dt_for_role, role, period.period_end_datetime)
#                 else:
#                     current_dt_for_role += timedelta(minutes=30)  # Small advancement
            
#             if show_this_iteration:
#                 print(f"  Next iteration starts at: {current_dt_for_role}")
            
#             if current_dt_for_role >= period.period_end_datetime:
#                 if show_this_iteration:
#                     print(f"  STOPPING: Reached period end")
#                 break
                
#         if iter_count >= max_iter: 
#             flash(f"Max iterations for role '{role.name}' during slot generation.", "warning")
        
#         total_new_slots_generated += role_slots_generated_this_role
        
#         # === FINAL ROLE SUMMARY ===
#         if should_debug_this_role:
#             print(f"\nFINAL SUMMARY FOR {role.name}:")
#             print(f"  - Total slots generated: {role_slots_generated_this_role}")
#             print(f"  - Total iterations: {iter_count}")
#             print(f"  - Slots per day (approx): {role_slots_generated_this_role / max(1, (period.period_end_datetime - period.period_start_datetime).days):.1f}")
#             if role.has_time_restrictions():
#                 working_hours_per_day = 8  # Approximate
#                 max_possible_slots_per_day = working_hours_per_day / max(1, duration.total_seconds() / 3600)
#                 print(f"  - Theoretical max slots/day: {max_possible_slots_per_day:.1f}")
#             print("=" * 60)
        
#         # Log information about what was generated
#         if role.has_time_restrictions():
#             current_app.logger.info(f"Generated {role_slots_generated_this_role} time-restricted slots for role '{role.name}' ({role.get_working_hours_str()})")
#         else:
#             current_app.logger.info(f"Generated {role_slots_generated_this_role} all-day slots for role '{role.name}'")
    
#     # === FINAL DEBUG SUMMARY ===
#     if DEBUG_SLOT_GENERATION:
#         print(f"\n{'='*80}")
#         print(f"FINAL PERIOD SUMMARY")
#         print(f"{'='*80}")
#         print(f"Total slots generated across all roles: {total_new_slots_generated}")
#         for role in job_roles_for_period:
#             role_count = sum(1 for slot in generated_slot_objects if slot.job_role_id == role.id)
#             print(f"  - {role.name}: {role_count} slots")
#         print(f"Random seed used: {random_seed}")
#         print("=" * 80)
    
#     if total_new_slots_generated > 0:
#         try:
#             db.session.commit()
#             flash(f"{total_new_slots_generated} coverage slots generated for '{period.name}'. Attempting assignment...", "info")
#             current_app.logger.info(f"{total_new_slots_generated} ShiftDefinition slots committed for period {period.id}.")
#         except Exception as e:
#             db.session.rollback()
#             flash(f"Error committing generated slots: {e}", "danger")
#             current_app.logger.error(f"Error committing slots for period {period.id}: {e}")
#             return redirect(url_for('main.manage_job_roles_for_period', period_id=period.id))
#     else:
#         flash("No new coverage slots were generated. Check role durations. No assignments will be made.", "warning")
#         return redirect(url_for('main.manage_job_roles_for_period', period_id=period.id))

#     # --- Step 3: Prepare for and run assignment algorithm ---
#     workers = Worker.query.options(selectinload(Worker.qualified_roles)).all()
#     if not workers:
#         flash("No workers found. Slots generated, but assignments cannot proceed.", "warning")
#         session['assignment_details'] = [("warning", "No workers found in the system to perform assignments.")]
#         return redirect(url_for('main.manage_job_roles_for_period', period_id=period.id))

#     # ============ RANDOMIZE WORKER ORDER ============
#     random.shuffle(workers)
#     current_app.logger.info(f"Randomized worker order for assignment ({len(workers)} workers)")

#     assignments_to_make = []
#     for slot_def in generated_slot_objects:
#         if slot_def.id is None: 
#             continue
#         assignments_to_make.append(ScheduledShift(shift_definition_id=slot_def.id))
    
#     if assignments_to_make:
#         db.session.add_all(assignments_to_make)
#         db.session.commit()
#         current_app.logger.info(f"{len(assignments_to_make)} ScheduledShift placeholders created for period {period.id}.")
#     else:
#         flash("No assignment placeholders created, though slots generated. Unexpected error.", "danger")
#         current_app.logger.error(f"Failed to create ScheduledShift placeholders for period {period.id} despite {total_new_slots_generated} slots.")
#         return redirect(url_for('main.manage_job_roles_for_period', period_id=period.id))

#     from .algorithm import assign_shifts_fairly
#     all_pending_assignments = ScheduledShift.query.options(
#             joinedload(ScheduledShift.defined_slot).joinedload(ShiftDefinition.job_role)
#         ).join(ShiftDefinition)\
#         .filter(ShiftDefinition.scheduling_period_id == period.id, ScheduledShift.worker_id.is_(None))\
#         .all()
    
#     # ============ RANDOMIZE ASSIGNMENT ORDER ============
#     random.shuffle(all_pending_assignments)
#     current_app.logger.info(f"Randomized assignment order for {len(all_pending_assignments)} pending assignments")
    
#     current_app.logger.info(f"Attempting to assign {len(all_pending_assignments)} slots for period {period.id}.")
#     assignment_successful, algo_messages_raw = assign_shifts_fairly(all_pending_assignments, workers, period)
    
#     # --- Message handling ---
#     detailed_assignment_warnings = []
#     error_count = 0
#     warning_summary_count = 0

#     for msg_type, msg_text in algo_messages_raw:
#         if msg_type == "error":
#             error_count += 1
#             current_app.logger.error(f"Algo Error: {msg_text}")
#             flash(f"Critical Algorithm Error: {msg_text}", "danger")
#         elif msg_type == "warning":
#             warning_summary_count += 1
#             detailed_assignment_warnings.append(msg_text)
#             current_app.logger.warning(f"Algo Warning: {msg_text}")
#         elif msg_type == "success":
#             flash(msg_text, "success")
#         else:
#             flash(msg_text, msg_type)
    
#     session['assignment_details'] = detailed_assignment_warnings

#     if error_count > 0:
#         flash(f"{error_count} critical errors occurred during assignment. Check server logs.", "danger")
    
#     if warning_summary_count > 0:
#         flash(f"Assignment complete: {warning_summary_count} slots could not be filled. See details below or check server logs.", "warning")
    
#     # ============ ADD SEED TO SUCCESS/INFO MESSAGES ============
#     if assignment_successful and error_count == 0 and warning_summary_count == 0:
#         flash(f"All shifts assigned successfully! (Random seed: {random_seed})", "success")
#     elif not assignment_successful and error_count == 0 and warning_summary_count == 0:
#         flash(f"Shift assignment process completed, but the algorithm reported not all shifts filled (Random seed: {random_seed})", "warning")
#     elif total_new_slots_generated > 0 and not all_pending_assignments and (error_count > 0 or warning_summary_count > 0):
#         flash(f"Slots were generated, but assignment step encountered issues before processing. (Random seed: {random_seed})", "danger")

#     return redirect(url_for('main.manage_job_roles_for_period', period_id=period.id))

# Replace the ENTIRE slot generation loop in generate_slots_and_assign_action
# Find the section that starts with "for role in job_roles_for_period:" and replace everything
# inside that loop until the end of the loop

@main_bp.route('/period/<int:period_id>/generate_slots_and_assign', methods=['POST'])
def generate_slots_and_assign_action(period_id):
    period = SchedulingPeriod.query.get_or_404(period_id)
    
    # ============ RANDOM SEED HANDLING ============
    random_seed = request.form.get('random_seed', '').strip()
    if random_seed:
        try:
            random_seed = int(random_seed)
            if random_seed < 1 or random_seed > 999999:
                flash("Random seed must be between 1 and 999999. Using auto-generated seed.", "warning")
                random_seed = random.randint(1, 999999)
        except ValueError:
            flash("Invalid random seed. Using auto-generated seed.", "warning")
            random_seed = random.randint(1, 999999)
    else:
        random_seed = random.randint(1, 999999)
    
    random.seed(random_seed)
    current_app.logger.info(f"Using random seed: {random_seed}")
    
    # === DEBUGGING CONFIGURATION ===
    DEBUG_SLOT_GENERATION = True
    DEBUG_ROLE_NAMES = []  # Empty = debug ALL roles, or specify: ["Toran", "Cook"]
    DEBUG_MAX_ITERATIONS_TO_SHOW = 10
    
    current_app.logger.info(f"Clearing old data for period {period.id} ('{period.name}')")
    ids_to_delete_assignments = [s.id for s in ScheduledShift.query.join(ShiftDefinition)
                               .filter(ShiftDefinition.scheduling_period_id == period_id).all()]
    if ids_to_delete_assignments:
        ScheduledShift.query.filter(ScheduledShift.id.in_(ids_to_delete_assignments)).delete(synchronize_session=False)
    ShiftDefinition.query.filter_by(scheduling_period_id=period.id).delete()
    db.session.commit()
    current_app.logger.info(f"Old data cleared for period {period.id}.")

    job_roles_for_period = JobRole.query.filter_by(scheduling_period_id=period.id).all()
    if not job_roles_for_period:
        flash("No job roles defined for this period. Cannot generate slots or assign.", "warning")
        return redirect(url_for('main.manage_job_roles_for_period', period_id=period.id))

    random.shuffle(job_roles_for_period)
    current_app.logger.info(f"Randomized processing order for {len(job_roles_for_period)} job roles")

    total_new_slots_generated = 0
    generated_slot_objects = []
    
    for role in job_roles_for_period:
        should_debug_this_role = (
            DEBUG_SLOT_GENERATION and 
            (not DEBUG_ROLE_NAMES or role.name in DEBUG_ROLE_NAMES)
        )
        
        if should_debug_this_role:
            print(f"\n{'='*60}")
            print(f"DEBUGGING ROLE: {role.name}")
            print(f"{'='*60}")
            print(f"Configuration:")
            print(f"  - Number needed: {role.number_needed}")
            print(f"  - Duration: {role.get_duration_timedelta()}")
            print(f"  - Has time restrictions: {role.has_time_restrictions()}")
            if role.has_time_restrictions():
                print(f"  - Work hours: {role.work_start_time} - {role.work_end_time}")
                print(f"  - Is overnight: {role.is_overnight_shift}")
            print(f"  - Period: {period.period_start_datetime} to {period.period_end_datetime}")
            print(f"  - Random seed: {random_seed}")
            print("-" * 60)
        
        role_slots_generated_this_role = 0
        current_dt_for_role = period.period_start_datetime
        duration = role.get_duration_timedelta()
        
        if duration.total_seconds() <= 0:
            current_app.logger.warning(f"Skipping role '{role.name}' due to zero duration for period {period.id}.")
            flash(f"Job Role '{role.name}' has zero/negative shift duration and was skipped for slot generation.", "warning")
            continue
        
        # If role has time restrictions, start at the first valid time
        if role.has_time_restrictions():
            if not is_time_within_role_restrictions(current_dt_for_role, role):
                current_dt_for_role = get_next_valid_start_time(current_dt_for_role, role, period.period_end_datetime)
                if should_debug_this_role:
                    print(f"Initial time not valid, jumping to first valid time: {current_dt_for_role}")
            
        max_iter = 5000
        iter_count = 0
        iterations_shown = 0
        consecutive_invalid_slots = 0
        
        while current_dt_for_role < period.period_end_datetime and iter_count < max_iter:
            iter_count += 1
            show_this_iteration = (
                should_debug_this_role and 
                iterations_shown < DEBUG_MAX_ITERATIONS_TO_SHOW
            )
            
            if show_this_iteration:
                print(f"\nIteration {iter_count}:")
                print(f"  Current time: {current_dt_for_role}")
            
            # Verify current time is within role's working hours
            if role.has_time_restrictions():
                is_valid_time = is_time_within_role_restrictions(current_dt_for_role, role)
                
                if show_this_iteration:
                    print(f"  Time restriction check: {is_valid_time}")
                    print(f"  Current time only: {current_dt_for_role.time()}")
                    print(f"  Work window: {role.work_start_time} - {role.work_end_time}")
                
                if not is_valid_time:
                    if show_this_iteration:
                        print(f"  SKIPPING: Time not within restrictions")
                    
                    current_dt_for_role = get_next_valid_start_time(current_dt_for_role, role, period.period_end_datetime)
                    consecutive_invalid_slots = 0
                    
                    if show_this_iteration:
                        print(f"  Moved to next valid time: {current_dt_for_role}")
                    continue
            
            # Calculate slot boundaries
            slot_start = current_dt_for_role
            slot_end = current_dt_for_role + duration
            
            if show_this_iteration:
                print(f"  Proposed slot: {slot_start} to {slot_end}")
            
            # Constrain to working hours if needed
            if role.has_time_restrictions():
                original_slot_end = slot_end
                slot_end = constrain_slot_to_working_hours(slot_start, slot_end, role)
                
                if show_this_iteration and slot_end != original_slot_end:
                    print(f"  Time constraint: Adjusted end from {original_slot_end} to {slot_end}")
            
            # Ensure slot doesn't exceed period end
            if slot_end > period.period_end_datetime:
                original_slot_end = slot_end
                slot_end = period.period_end_datetime
                if show_this_iteration:
                    print(f"  Period limit: Adjusted end from {original_slot_end} to {slot_end}")
            
            # Create slot if valid
            if slot_start < slot_end:
                if show_this_iteration:
                    print(f"  ✓ CREATING {role.number_needed} slot(s) from {slot_start} to {slot_end}")
                    print(f"    Duration: {(slot_end - slot_start).total_seconds() / 3600:.2f} hours")
                
                for i in range(1, role.number_needed + 1):
                    new_slot = ShiftDefinition(
                        slot_start_datetime=slot_start,
                        slot_end_datetime=slot_end,
                        instance_number=i,
                        scheduling_period_id=period.id,
                        job_role_id=role.id
                    )
                    db.session.add(new_slot)
                    generated_slot_objects.append(new_slot)
                    role_slots_generated_this_role += 1
                    
                    if show_this_iteration:
                        print(f"    - Created slot #{i}: {new_slot.name}")
                
                if show_this_iteration:
                    iterations_shown += 1
                
                # Calculate next start time
                tentative_next_start = slot_start + duration
                
                # Check if next start is valid
                if role.has_time_restrictions():
                    if not is_time_within_role_restrictions(tentative_next_start, role):
                        # Jump to next valid start time
                        current_dt_for_role = get_next_valid_start_time(
                            tentative_next_start, role, period.period_end_datetime
                        )
                        if show_this_iteration:
                            print(f"  Next start {tentative_next_start} invalid, jumping to {current_dt_for_role}")
                    else:
                        current_dt_for_role = tentative_next_start
                        if show_this_iteration:
                            print(f"  Next start: {current_dt_for_role}")
                else:
                    current_dt_for_role = tentative_next_start
                    if show_this_iteration:
                        print(f"  Next start: {current_dt_for_role}")
                
                consecutive_invalid_slots = 0
                
            else:
                # Invalid slot
                consecutive_invalid_slots += 1
                
                if show_this_iteration:
                    print(f"  ✗ INVALID SLOT: start >= end ({slot_start} >= {slot_end})")
                    print(f"    Consecutive invalid: {consecutive_invalid_slots}")
                
                # Prevent infinite loop
                if consecutive_invalid_slots >= 10:
                    if show_this_iteration:
                        print(f"  BREAKING: Too many consecutive invalid slots")
                    
                    # Force advance to next valid time
                    if role.has_time_restrictions():
                        current_dt_for_role = get_next_valid_start_time(
                            current_dt_for_role + timedelta(hours=1),
                            role,
                            period.period_end_datetime
                        )
                    else:
                        current_dt_for_role += timedelta(hours=1)
                    
                    consecutive_invalid_slots = 0
                    continue
                
                # Try smaller advancement
                if role.has_time_restrictions():
                    current_dt_for_role = get_next_valid_start_time(
                        current_dt_for_role,
                        role,
                        period.period_end_datetime
                    )
                else:
                    current_dt_for_role += timedelta(minutes=30)
            
            if current_dt_for_role >= period.period_end_datetime:
                if show_this_iteration:
                    print(f"  STOPPING: Reached period end")
                break
        
        if iter_count >= max_iter:
            flash(f"Max iterations for role '{role.name}' during slot generation.", "warning")
        
        total_new_slots_generated += role_slots_generated_this_role
        
        if should_debug_this_role:
            print(f"\nFINAL SUMMARY FOR {role.name}:")
            print(f"  - Total slots generated: {role_slots_generated_this_role}")
            print(f"  - Total iterations: {iter_count}")
            period_days = (period.period_end_datetime - period.period_start_datetime).days
            if period_days > 0:
                print(f"  - Slots per day (approx): {role_slots_generated_this_role / period_days:.1f}")
            if role.has_time_restrictions():
                working_hours_per_day = (
                    (datetime.combine(date.today(), role.work_end_time) -
                     datetime.combine(date.today(), role.work_start_time)).total_seconds() / 3600
                )
                if role.is_overnight_shift:
                    working_hours_per_day = 24 - working_hours_per_day
                max_possible_slots_per_day = working_hours_per_day / max(1, duration.total_seconds() / 3600)
                print(f"  - Working hours per day: {working_hours_per_day:.1f}h")
                print(f"  - Theoretical max slots/day: {max_possible_slots_per_day:.1f}")
            print("=" * 60)
        
        if role.has_time_restrictions():
            current_app.logger.info(f"Generated {role_slots_generated_this_role} time-restricted slots for role '{role.name}' ({role.get_working_hours_str()})")
        else:
            current_app.logger.info(f"Generated {role_slots_generated_this_role} all-day slots for role '{role.name}'")
    
    if DEBUG_SLOT_GENERATION:
        print(f"\n{'='*80}")
        print(f"FINAL PERIOD SUMMARY")
        print(f"{'='*80}")
        print(f"Total slots generated across all roles: {total_new_slots_generated}")
        for role in job_roles_for_period:
            role_count = sum(1 for slot in generated_slot_objects if slot.job_role_id == role.id)
            print(f"  - {role.name}: {role_count} slots")
        print(f"Random seed used: {random_seed}")
        print("=" * 80)
    
    if total_new_slots_generated > 0:
        try:
            db.session.commit()
            flash(f"{total_new_slots_generated} coverage slots generated for '{period.name}'. Attempting assignment...", "info")
            current_app.logger.info(f"{total_new_slots_generated} ShiftDefinition slots committed for period {period.id}.")
        except Exception as e:
            db.session.rollback()
            flash(f"Error committing generated slots: {e}", "danger")
            current_app.logger.error(f"Error committing slots for period {period.id}: {e}")
            return redirect(url_for('main.manage_job_roles_for_period', period_id=period.id))
    else:
        flash("No new coverage slots were generated. Check role durations. No assignments will be made.", "warning")
        return redirect(url_for('main.manage_job_roles_for_period', period_id=period.id))

    # --- Step 3: Prepare for and run assignment algorithm ---
    workers = Worker.query.options(selectinload(Worker.qualified_roles)).all()
    if not workers:
        flash("No workers found. Slots generated, but assignments cannot proceed.", "warning")
        session['assignment_details'] = [("warning", "No workers found in the system to perform assignments.")]
        return redirect(url_for('main.manage_job_roles_for_period', period_id=period.id))

    random.shuffle(workers)
    current_app.logger.info(f"Randomized worker order for assignment ({len(workers)} workers)")

    assignments_to_make = []
    for slot_def in generated_slot_objects:
        if slot_def.id is None:
            continue
        assignments_to_make.append(ScheduledShift(shift_definition_id=slot_def.id))
    
    if assignments_to_make:
        db.session.add_all(assignments_to_make)
        db.session.commit()
        current_app.logger.info(f"{len(assignments_to_make)} ScheduledShift placeholders created for period {period.id}.")
    else:
        flash("No assignment placeholders created, though slots generated. Unexpected error.", "danger")
        current_app.logger.error(f"Failed to create ScheduledShift placeholders for period {period.id} despite {total_new_slots_generated} slots.")
        return redirect(url_for('main.manage_job_roles_for_period', period_id=period.id))

    from .algorithm import assign_shifts_fairly
    all_pending_assignments = ScheduledShift.query.options(
            joinedload(ScheduledShift.defined_slot).joinedload(ShiftDefinition.job_role)
        ).join(ShiftDefinition)\
        .filter(ShiftDefinition.scheduling_period_id == period.id, ScheduledShift.worker_id.is_(None))\
        .all()
    
    random.shuffle(all_pending_assignments)
    current_app.logger.info(f"Randomized assignment order for {len(all_pending_assignments)} pending assignments")
    
    current_app.logger.info(f"Attempting to assign {len(all_pending_assignments)} slots for period {period.id}.")
    assignment_successful, algo_messages_raw = assign_shifts_fairly(all_pending_assignments, workers, period)
    
    # --- Message handling ---
    detailed_assignment_warnings = []
    error_count = 0
    warning_summary_count = 0

    for msg_type, msg_text in algo_messages_raw:
        if msg_type == "error":
            error_count += 1
            current_app.logger.error(f"Algo Error: {msg_text}")
            flash(f"Critical Algorithm Error: {msg_text}", "danger")
        elif msg_type == "warning":
            warning_summary_count += 1
            detailed_assignment_warnings.append(msg_text)
            current_app.logger.warning(f"Algo Warning: {msg_text}")
        elif msg_type == "success":
            flash(msg_text, "success")
        else:
            flash(msg_text, msg_type)
    
    session['assignment_details'] = detailed_assignment_warnings

    if error_count > 0:
        flash(f"{error_count} critical errors occurred during assignment. Check server logs.", "danger")
    
    if warning_summary_count > 0:
        flash(f"Assignment complete: {warning_summary_count} slots could not be filled. See details below or check server logs.", "warning")
    
    if assignment_successful and error_count == 0 and warning_summary_count == 0:
        flash(f"All shifts assigned successfully! (Random seed: {random_seed})", "success")
    elif not assignment_successful and error_count == 0 and warning_summary_count == 0:
        flash(f"Shift assignment process completed, but the algorithm reported not all shifts filled (Random seed: {random_seed})", "warning")
    elif total_new_slots_generated > 0 and not all_pending_assignments and (error_count > 0 or warning_summary_count > 0):
        flash(f"Slots were generated, but assignment step encountered issues before processing. (Random seed: {random_seed})", "danger")

    return redirect(url_for('main.manage_job_roles_for_period', period_id=period.id))


# Replace the get_next_valid_start_time function
def get_next_valid_start_time(current_time, role, period_end):
    """Get the next valid start time for a role with time restrictions"""
    if not role.has_time_restrictions():
        return current_time + timedelta(minutes=30)  # Default advancement
    
    current_time_only = current_time.time()
    
    if role.is_overnight_shift:
        # For overnight shifts (e.g., 22:00 - 06:00)
        if current_time_only >= role.work_start_time:
            # Already in the valid window (22:00-23:59), move to next day's start
            next_start = current_time.replace(
                hour=role.work_start_time.hour,
                minute=role.work_start_time.minute,
                second=0,
                microsecond=0
            ) + timedelta(days=1)
        elif current_time_only < role.work_end_time:
            # In the early morning valid window (00:00-06:00), move to today's start time
            next_start = current_time.replace(
                hour=role.work_start_time.hour,
                minute=role.work_start_time.minute,
                second=0,
                microsecond=0
            )
        else:
            # Between end and start (06:00-22:00), move to today's start time
            next_start = current_time.replace(
                hour=role.work_start_time.hour,
                minute=role.work_start_time.minute,
                second=0,
                microsecond=0
            )
        
        return min(next_start, period_end)
    else:
        # For day shifts (e.g., 08:00 - 22:00)
        if current_time_only < role.work_start_time:
            # Before start time today - move to start time today
            return current_time.replace(
                hour=role.work_start_time.hour,
                minute=role.work_start_time.minute,
                second=0,
                microsecond=0
            )
        else:
            # After start time today - move to start time tomorrow
            next_start = current_time.replace(
                hour=role.work_start_time.hour,
                minute=role.work_start_time.minute,
                second=0,
                microsecond=0
            ) + timedelta(days=1)
            return min(next_start, period_end)


# Replace the constrain_slot_to_working_hours function
def constrain_slot_to_working_hours(slot_start, slot_end, role):
    """Constrain a slot's end time to role's working hours"""
    if not role.has_time_restrictions():
        return slot_end
    
    start_time_only = slot_start.time()
    
    if role.is_overnight_shift:
        # For overnight shifts spanning midnight (e.g., 22:00 - 06:00)
        
        if start_time_only >= role.work_start_time:
            # Started on the "late" side (e.g., 22:00-23:59)
            # End time is next day at work_end_time
            working_hours_end = slot_start.replace(
                hour=role.work_end_time.hour,
                minute=role.work_end_time.minute,
                second=0,
                microsecond=0
            ) + timedelta(days=1)
        elif start_time_only < role.work_end_time:
            # Started on the "early" side (e.g., 00:00-05:59)
            # End time is same day at work_end_time
            working_hours_end = slot_start.replace(
                hour=role.work_end_time.hour,
                minute=role.work_end_time.minute,
                second=0,
                microsecond=0
            )
        else:
            # This shouldn't happen if validation is working correctly
            # Fall back to end of today
            working_hours_end = slot_start.replace(
                hour=role.work_end_time.hour,
                minute=role.work_end_time.minute,
                second=0,
                microsecond=0
            )
        
        return min(slot_end, working_hours_end)
    else:
        # For same-day shifts (e.g., 08:00 - 22:00)
        same_day_end = slot_start.replace(
            hour=role.work_end_time.hour,
            minute=role.work_end_time.minute,
            second=0,
            microsecond=0
        )
        
        # If the end time is before or at the start time, it means we've crossed midnight
        # (shouldn't happen for non-overnight shifts, but handle it)
        if same_day_end <= slot_start:
            same_day_end += timedelta(days=1)
        
        return min(slot_end, same_day_end)



# Add these new export routes at the end of your routes.py file, before the final comment
@main_bp.route('/period/<int:period_id>/export_schedule_csv')
def export_schedule_csv(period_id):
    """Export the schedule for a specific period as CSV"""
    period = SchedulingPeriod.query.get_or_404(period_id)
    
    # Get all scheduled assignments for this period
    scheduled_assignments = ScheduledShift.query.options(
        joinedload(ScheduledShift.defined_slot).joinedload(ShiftDefinition.job_role),
        joinedload(ScheduledShift.worker_assigned)
    ).join(ShiftDefinition).\
    filter(ShiftDefinition.scheduling_period_id == period.id).\
    order_by(ShiftDefinition.slot_start_datetime, ShiftDefinition.job_role_id, ShiftDefinition.instance_number).all()
    
    # Create CSV data
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow([
        'Role & Instance',
        'Start Date',
        'Start Time', 
        'End Date',
        'End Time',
        'Duration',
        'Assigned Worker',
        'Worker Email'
    ])
    
    # Write data rows
    for assignment in scheduled_assignments:
        slot = assignment.defined_slot
        worker = assignment.worker_assigned
        
        writer.writerow([
            slot.name,
            slot.slot_start_datetime.strftime('%Y-%m-%d'),
            slot.slot_start_datetime.strftime('%H:%M'),
            slot.slot_end_datetime.strftime('%Y-%m-%d'), 
            slot.slot_end_datetime.strftime('%H:%M'),
            slot.duration_hours_minutes_str,
            worker.name if worker else 'UNASSIGNED',
            worker.email if worker and worker.email else ''
        ])
    
    # Create response
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = f'attachment; filename=schedule_{period.name.replace(" ", "_")}_{datetime.now().strftime("%Y%m%d")}.csv'
    
    return response

@main_bp.route('/period/<int:period_id>/export_schedule_excel')
def export_schedule_excel(period_id):
    """Export the schedule for a specific period as Excel"""
    period = SchedulingPeriod.query.get_or_404(period_id)
    
    # Get all scheduled assignments for this period
    scheduled_assignments = ScheduledShift.query.options(
        joinedload(ScheduledShift.defined_slot).joinedload(ShiftDefinition.job_role),
        joinedload(ScheduledShift.worker_assigned)
    ).join(ShiftDefinition).\
    filter(ShiftDefinition.scheduling_period_id == period.id).\
    order_by(ShiftDefinition.slot_start_datetime, ShiftDefinition.job_role_id, ShiftDefinition.instance_number).all()
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    unassigned_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add title
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f"Schedule for {period.name}"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Add period info
    ws.merge_cells('A2:H2')
    period_cell = ws['A2']
    period_cell.value = f"Period: {period.period_start_datetime.strftime('%Y-%m-%d %H:%M')} to {period.period_end_datetime.strftime('%Y-%m-%d %H:%M')}"
    period_cell.alignment = Alignment(horizontal='center')
    
    # Add headers starting from row 4
    headers = [
        'Role & Instance',
        'Start Date',
        'Start Time',
        'End Date', 
        'End Time',
        'Duration',
        'Assigned Worker',
        'Worker Email'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Add data starting from row 5
    for row, assignment in enumerate(scheduled_assignments, 5):
        slot = assignment.defined_slot
        worker = assignment.worker_assigned
        
        data = [
            slot.name,
            slot.slot_start_datetime.strftime('%Y-%m-%d'),
            slot.slot_start_datetime.strftime('%H:%M'),
            slot.slot_end_datetime.strftime('%Y-%m-%d'),
            slot.slot_end_datetime.strftime('%H:%M'),
            slot.duration_hours_minutes_str,
            worker.name if worker else 'UNASSIGNED',
            worker.email if worker and worker.email else ''
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            
            # Highlight unassigned shifts
            if not worker and col <= 8:
                cell.fill = unassigned_fill
    
    # Auto-adjust column widths
    from openpyxl.utils import get_column_letter
    
    for col_num in range(1, 9):  # We have 8 columns (A to H)
        max_length = 0
        for row in ws.iter_rows(min_col=col_num, max_col=col_num):
            for cell in row:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = min(max_length + 2, 50)
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create response
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=schedule_{period.name.replace(" ", "_")}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    
    return response


# Import the function here to avoid import issues
from .algorithm import is_worker_qualified_for_slot

# For editing time periods, you can add a route like this:
@main_bp.route('/period/<int:period_id>/edit', methods=['GET', 'POST'])
def edit_period(period_id):
    period = SchedulingPeriod.query.get_or_404(period_id)
    
    if request.method == 'POST':
        try:
            name = request.form.get('period_name')
            start_date_str = request.form.get('period_start_date_hidden')
            end_date_str = request.form.get('period_end_date_hidden')
            start_time_str = request.form.get('period_start_time')
            end_time_str = request.form.get('period_end_time')

            if not name or not name.strip():
                flash("Period name is required.", "danger")
                return redirect(url_for('main.edit_period', period_id=period_id))
            name = name.strip()

            if not all([start_date_str, end_date_str, start_time_str, end_time_str]):
                flash("All period date and time fields are required.", "danger")
                return redirect(url_for('main.edit_period', period_id=period_id))

            # Check if name is taken by another period (not this one)
            existing_period = SchedulingPeriod.query.filter(
                SchedulingPeriod.name.ilike(name),
                SchedulingPeriod.id != period_id
            ).first()
            if existing_period:
                flash(f"A scheduling period with the name '{name}' already exists.", "danger")
                return redirect(url_for('main.edit_period', period_id=period_id))

            start_date_obj = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date_obj = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            start_time_obj = datetime.strptime(start_time_str, '%H:%M').time()
            end_time_obj = datetime.strptime(end_time_str, '%H:%M').time()

            if end_date_obj < start_date_obj or \
               (end_date_obj == start_date_obj and end_time_obj <= start_time_obj):
                flash("Period end must be after period start.", "danger")
                return redirect(url_for('main.edit_period', period_id=period_id))

            period_start_dt = datetime.combine(start_date_obj, start_time_obj)
            period_end_dt = datetime.combine(end_date_obj, end_time_obj)
            
            # Update the period
            period.name = name
            period.period_start_datetime = period_start_dt
            period.period_end_datetime = period_end_dt

            db.session.commit()
            flash(f"Scheduling Period '{name}' updated successfully.", "success")
            return redirect(url_for('main.manage_periods'))
            
        except ValueError as ve:
            flash(f"Invalid date or time format: {ve}", "danger")
        except Exception as e:
            db.session.rollback()
            flash(f"Error updating period: {e}", "danger")
            current_app.logger.error(f"Error updating period {period_id}: {e}\n{request.form}")
        return redirect(url_for('main.edit_period', period_id=period_id))

    # GET request - show edit form
    return render_template('edit_period.html', period=period)


    # Option to edit the "Home\Dashboard" list manually if needed
    # Add these routes to your routes.py file for manual assignment editing

@main_bp.route('/assignment/<int:assignment_id>/edit_worker', methods=['POST'])
def edit_assignment_worker(assignment_id):
    """Manually assign or reassign a worker to a specific shift"""
    assignment = ScheduledShift.query.get_or_404(assignment_id)
    new_worker_id = request.form.get('worker_id')
    
    # Get the shift definition to check role requirements
    shift_def = assignment.defined_slot
    if not shift_def:
        flash("Error: Shift definition not found.", "danger")
        return redirect(url_for('main.index'))
    
    try:
        if new_worker_id == 'unassign':
            # Unassign the worker
            old_worker_name = assignment.worker_assigned.name if assignment.worker_assigned else "unassigned"
            assignment.worker_id = None
            db.session.commit()
            flash(f"Unassigned {old_worker_name} from {shift_def.name} on {shift_def.slot_start_datetime.strftime('%Y-%m-%d %H:%M')}", "info")
            
        elif new_worker_id and new_worker_id.isdigit():
            # Assign to a specific worker
            new_worker = Worker.query.get(int(new_worker_id))
            if not new_worker:
                flash("Worker not found.", "danger")
                return redirect(url_for('main.index'))
            
            # Check if worker is qualified for this role
            if not is_worker_qualified_for_slot(new_worker, shift_def):
                flash(f"Warning: {new_worker.name} is not qualified for role {shift_def.job_role.name}, but assignment was made anyway.", "warning")
            
            # Check for conflicts with other assignments
            active_period = get_active_period()
            if active_period:
                # Get all other assignments for this worker in the same period
                conflicting_assignments = ScheduledShift.query.options(
                    joinedload(ScheduledShift.defined_slot)
                ).join(ShiftDefinition).filter(
                    ShiftDefinition.scheduling_period_id == active_period.id,
                    ScheduledShift.worker_id == new_worker.id,
                    ScheduledShift.id != assignment_id
                ).all()
                
                # Check for time overlaps
                for other_assignment in conflicting_assignments:
                    other_slot = other_assignment.defined_slot
                    if (max(shift_def.slot_start_datetime, other_slot.slot_start_datetime) < 
                        min(shift_def.slot_end_datetime, other_slot.slot_end_datetime)):
                        flash(f"Warning: Time conflict detected with {other_slot.name} on {other_slot.slot_start_datetime.strftime('%Y-%m-%d %H:%M')}, but assignment was made anyway.", "warning")
                        break
            
            old_worker_name = assignment.worker_assigned.name if assignment.worker_assigned else "unassigned"
            assignment.worker_id = new_worker.id
            db.session.commit()
            
            if old_worker_name != "unassigned":
                flash(f"Reassigned {shift_def.name} from {old_worker_name} to {new_worker.name} on {shift_def.slot_start_datetime.strftime('%Y-%m-%d %H:%M')}", "success")
            else:
                flash(f"Assigned {new_worker.name} to {shift_def.name} on {shift_def.slot_start_datetime.strftime('%Y-%m-%d %H:%M')}", "success")
        else:
            flash("Invalid worker selection.", "danger")
            
    except Exception as e:
        db.session.rollback()
        flash(f"Error updating assignment: {e}", "danger")
        current_app.logger.error(f"Error in edit_assignment_worker for assignment {assignment_id}: {e}")
    
    return redirect(url_for('main.index'))

@main_bp.route('/assignment/<int:assignment_id>/swap', methods=['POST'])
def swap_assignments(assignment_id):
    """Swap workers between two assignments"""
    assignment1 = ScheduledShift.query.get_or_404(assignment_id)
    assignment2_id = request.form.get('swap_with_assignment_id')
    
    if not assignment2_id or not assignment2_id.isdigit():
        flash("Invalid assignment to swap with.", "danger")
        return redirect(url_for('main.index'))
    
    assignment2 = ScheduledShift.query.get(int(assignment2_id))
    if not assignment2:
        flash("Assignment to swap with not found.", "danger")
        return redirect(url_for('main.index'))
    
    try:
        # Get worker info before swap
        worker1 = assignment1.worker_assigned
        worker2 = assignment2.worker_assigned
        shift1 = assignment1.defined_slot
        shift2 = assignment2.defined_slot
        
        # Perform the swap
        assignment1.worker_id = worker2.id if worker2 else None
        assignment2.worker_id = worker1.id if worker1 else None
        
        db.session.commit()
        
        worker1_name = worker1.name if worker1 else "unassigned"
        worker2_name = worker2.name if worker2 else "unassigned"
        
        flash(f"Swapped assignments: {worker1_name} ↔ {worker2_name} between {shift1.name} and {shift2.name}", "success")
        
    except Exception as e:
        db.session.rollback()
        flash(f"Error swapping assignments: {e}", "danger")
        current_app.logger.error(f"Error in swap_assignments: {e}")
    
    return redirect(url_for('main.index'))


    # Adding statistics page to know how many shifts were assigned, unassigned, how did shift at night, etc.
    # Also we want to see from each role how each worker performed, how many shifts were assigned to each worker, etc.

# The statistics route to display fairness and workload distribution
# from collections import defaultdict

# @main_bp.route('/period/<int:period_id>/fairness_statistics')
# def fairness_statistics(period_id):
#     """Display fairness and workload distribution statistics for a period."""
#     period = SchedulingPeriod.query.get_or_404(period_id)
#     all_workers = Worker.query.order_by(Worker.name).all()
#     all_roles = JobRole.query.filter_by(scheduling_period_id=period.id).order_by(JobRole.name).all()

#     if not all_workers:
#         flash("No workers found to generate statistics.", "warning")
#         return redirect(url_for('main.manage_workers'))

#     # Eagerly load all necessary data
#     all_assignments = ScheduledShift.query.options(
#         joinedload(ScheduledShift.defined_slot).joinedload(ShiftDefinition.job_role),
#         joinedload(ScheduledShift.worker_assigned)
#     ).join(ShiftDefinition).filter(
#         ShiftDefinition.scheduling_period_id == period.id
#     ).all()

#     # --- Initialize Statistics Dictionaries ---
#     stats = {
#         'total_shifts': len(all_assignments), 'assigned_shifts': 0, 'unassigned_shifts': 0,
#         'unassigned_shifts_list': [], 'worker_stats': {}, 'role_stats': {}
#     }
#     period_duration_hours = (period.period_end_datetime - period.period_start_datetime).total_seconds() / 3600.0

#     for worker in all_workers:
#         stats['worker_stats'][worker.id] = {
#             'name': worker.name, 'total_shifts': 0, 'total_hours': 0.0, 'night_shifts': 0,
#             'weekend_shifts': 0, 'difficulty_counts': defaultdict(int),
#             'downtime_hours': period_duration_hours, 'role_distribution': defaultdict(int)
#         }
#     for role in all_roles:
#         stats['role_stats'][role.id] = {
#             'name': role.name, 'total_shifts': 0, 'total_hours': 0.0,
#             'worker_distribution': defaultdict(int), 'unique_workers': 0
#         }

#     # --- Process All Assignments ---
#     NIGHT_START_TIME = time(0, 0)
#     NIGHT_END_TIME = time(6, 0)

#     for assignment in all_assignments:
#         slot = assignment.defined_slot
#         if not slot or not slot.job_role: continue
#         role = slot.job_role
#         stats['role_stats'][role.id]['total_shifts'] += 1
#         stats['role_stats'][role.id]['total_hours'] += slot.duration_total_seconds / 3600.0

#         if assignment.worker_assigned:
#             stats['assigned_shifts'] += 1
#             worker = assignment.worker_assigned
#             worker_stat = stats['worker_stats'][worker.id]
#             duration_hours = slot.duration_total_seconds / 3600.0
            
#             worker_stat['total_shifts'] += 1
#             worker_stat['total_hours'] += duration_hours
#             worker_stat['downtime_hours'] -= duration_hours
            
#             # Check for night shift
#             if slot.slot_start_datetime.time() < NIGHT_END_TIME:
#                 worker_stat['night_shifts'] += 1
            
#             # Check for weekend shift (Saturday is 5, Sunday is 6)
#             if slot.slot_start_datetime.weekday() >= 5:
#                 worker_stat['weekend_shifts'] += 1

#             # Difficulty & Role Distribution
#             worker_stat['difficulty_counts'][int(role.difficulty_multiplier)] += 1
#             worker_stat['role_distribution'][role.name] += 1
#             stats['role_stats'][role.id]['worker_distribution'][worker.name] += 1

#         else:
#             stats['unassigned_shifts'] += 1
#             stats['unassigned_shifts_list'].append(assignment)

#     for role_stat in stats['role_stats'].values():
#         role_stat['unique_workers'] = len(role_stat['worker_distribution'])

#     # --- Generate Key Fairness Insights ---
#     insights = []
#     if stats['assigned_shifts'] > 0:
#         active_workers = [w for w in stats['worker_stats'].values() if w['total_shifts'] > 0]
#         if not active_workers: active_workers = stats['worker_stats'].values()
        
#         avg_hours = sum(w['total_hours'] for w in active_workers) / len(active_workers)
#         avg_night = sum(w['night_shifts'] for w in active_workers) / len(active_workers)
#         avg_weekend = sum(w['weekend_shifts'] for w in active_workers) / len(active_workers)
        
#         for w in stats['worker_stats'].values():
#             if w['total_hours'] > avg_hours * 1.5:
#                 insights.append(f"<b>{w['name']}</b> is working significantly more hours ({w['total_hours']:.1f}) than the average ({avg_hours:.1f}).")
#             if w['total_hours'] < avg_hours * 0.5 and w['total_shifts'] > 0:
#                  insights.append(f"<b>{w['name']}</b> is working significantly fewer hours ({w['total_hours']:.1f}) than the average ({avg_hours:.1f}).")
#             if avg_night > 0.5 and w['night_shifts'] > avg_night * 1.75:
#                 insights.append(f"<b>{w['name']}</b> has a high number of night shifts ({w['night_shifts']}).")
#             if avg_weekend > 0.5 and w['weekend_shifts'] > avg_weekend * 1.75:
#                  insights.append(f"<b>{w['name']}</b> has a high number of weekend shifts ({w['weekend_shifts']}).")
#             if w['total_shifts'] == 0:
#                 insights.append(f"<b>{w['name']}</b> was not assigned any shifts.")
#     if stats['unassigned_shifts'] > 0:
#         insights.append(f"There are <b>{stats['unassigned_shifts']} unassigned shifts</b> that need coverage.")

#     return render_template('fairness_statistics.html', period=period, stats=stats, insights=insights)








# Enhanced fairness statistics route - REPLACE your existing fairness_statistics function with this
import math
from collections import defaultdict


@main_bp.route('/period/<int:period_id>/fairness_statistics')
def fairness_statistics(period_id):
    """Display enhanced fairness and workload distribution statistics for a period."""
    period = SchedulingPeriod.query.get_or_404(period_id)
    all_workers = Worker.query.order_by(Worker.name).all()
    all_roles = JobRole.query.filter_by(scheduling_period_id=period.id).order_by(JobRole.name).all()

    if not all_workers:
        flash("No workers found to generate statistics.", "warning")
        return redirect(url_for('main.manage_workers'))

    # Eagerly load all necessary data
    all_assignments = ScheduledShift.query.options(
        joinedload(ScheduledShift.defined_slot).joinedload(ShiftDefinition.job_role),
        joinedload(ScheduledShift.worker_assigned)
    ).join(ShiftDefinition).filter(
        ShiftDefinition.scheduling_period_id == period.id
    ).all()

    # --- Initialize Enhanced Statistics ---
    stats = {
        'total_shifts': len(all_assignments), 'assigned_shifts': 0, 'unassigned_shifts': 0,
        'unassigned_shifts_list': [], 'worker_stats': {}, 'role_stats': {},
        'fairness_metrics': {}, 'satisfaction_metrics': {}, 'envy_metrics': {}
    }
    
    period_duration_hours = (period.period_end_datetime - period.period_start_datetime).total_seconds() / 3600.0
    NIGHT_START_TIME = time(0, 0)
    NIGHT_END_TIME = time(6, 0)

    # Initialize worker stats with enhanced metrics
    for worker in all_workers:
        stats['worker_stats'][worker.id] = {
            'name': worker.name, 'total_shifts': 0, 'total_hours': 0.0, 'night_shifts': 0,
            'weekend_shifts': 0, 'difficulty_counts': defaultdict(int),
            'downtime_hours': period_duration_hours, 'role_distribution': defaultdict(int),
            'weighted_difficulty_burden': 0.0, 'personal_satisfaction_score': 0.0,
            'qualified_roles': [role.id for role in worker.qualified_roles if role.scheduling_period_id == period.id],
            'assignments': [], 'rest_periods': [], 'max_continuous_work': 0.0
        }
    
    for role in all_roles:
        stats['role_stats'][role.id] = {
            'name': role.name, 'total_shifts': 0, 'total_hours': 0.0,
            'worker_distribution': defaultdict(int), 'unique_workers': 0,
            'difficulty_rating': role.difficulty_multiplier
        }

    # --- Process All Assignments with Enhanced Tracking ---
    for assignment in all_assignments:
        slot = assignment.defined_slot
        if not slot or not slot.job_role: continue
        
        role = slot.job_role
        stats['role_stats'][role.id]['total_shifts'] += 1
        stats['role_stats'][role.id]['total_hours'] += slot.duration_total_seconds / 3600.0

        if assignment.worker_assigned:
            stats['assigned_shifts'] += 1
            worker = assignment.worker_assigned
            worker_stat = stats['worker_stats'][worker.id]
            duration_hours = slot.duration_total_seconds / 3600.0
            
            # Basic stats
            worker_stat['total_shifts'] += 1
            worker_stat['total_hours'] += duration_hours
            worker_stat['downtime_hours'] -= duration_hours
            
            # Store assignment details for advanced analysis
            worker_stat['assignments'].append({
                'start': slot.slot_start_datetime,
                'end': slot.slot_end_datetime,
                'duration': duration_hours,
                'role_id': role.id,
                'role_name': role.name,
                'difficulty': role.difficulty_multiplier,
                'is_night': slot.slot_start_datetime.time() < NIGHT_END_TIME or slot.slot_start_datetime.time() >= time(22, 0),
                'is_weekend': slot.slot_start_datetime.weekday() >= 5
            })
            
            # Night and weekend tracking
            if slot.slot_start_datetime.time() < NIGHT_END_TIME or slot.slot_start_datetime.time() >= time(22, 0):
                worker_stat['night_shifts'] += 1
            
            if slot.slot_start_datetime.weekday() >= 5:
                worker_stat['weekend_shifts'] += 1

            # Difficulty distribution and weighted burden
            difficulty_level = int(role.difficulty_multiplier)
            worker_stat['difficulty_counts'][difficulty_level] += 1
            worker_stat['weighted_difficulty_burden'] += duration_hours * role.difficulty_multiplier
            worker_stat['role_distribution'][role.name] += 1
            stats['role_stats'][role.id]['worker_distribution'][worker.name] += 1

        else:
            stats['unassigned_shifts'] += 1
            stats['unassigned_shifts_list'].append(assignment)

    # Calculate unique workers per role
    for role_stat in stats['role_stats'].values():
        role_stat['unique_workers'] = len(role_stat['worker_distribution'])

    # --- ENHANCED FAIRNESS METRICS CALCULATION ---
    
    # 1. WORKER SATISFACTION SCORES
    # Based on how "easy" their assigned work is relative to their own difficulty ratings
    for worker_id, worker_stat in stats['worker_stats'].items():
        if worker_stat['total_hours'] > 0:
            # Personal satisfaction = inverse of weighted difficulty burden
            avg_personal_difficulty = worker_stat['weighted_difficulty_burden'] / worker_stat['total_hours']
            # Higher satisfaction when assigned easier work (lower difficulty ratings)
            worker_stat['personal_satisfaction_score'] = max(0, (6 - avg_personal_difficulty) / 5) * 100
        else:
            worker_stat['personal_satisfaction_score'] = 50  # Neutral for unassigned workers

    # 2. PROPORTIONAL FAIRNESS
    # Each worker should get work proportional to their qualifications
    total_qualified_roles = sum(len(worker_stat['qualified_roles']) for worker_stat in stats['worker_stats'].values())
    if total_qualified_roles > 0:
        for worker_id, worker_stat in stats['worker_stats'].items():
            qualification_ratio = len(worker_stat['qualified_roles']) / total_qualified_roles if total_qualified_roles > 0 else 0
            expected_hours = qualification_ratio * sum(w['total_hours'] for w in stats['worker_stats'].values())
            actual_hours = worker_stat['total_hours']
            
            # Proportional fairness score (100 = perfectly proportional)
            if expected_hours > 0:
                proportional_score = min(100, (actual_hours / expected_hours) * 100)
            else:
                proportional_score = 100 if actual_hours == 0 else 0
            
            worker_stat['proportional_fairness_score'] = proportional_score

    # 3. ENVY-FREE ALLOCATION ANALYSIS
    envy_matrix = {}
    for worker_id1, worker_stat1 in stats['worker_stats'].items():
        envy_matrix[worker_id1] = {}
        worker1_satisfaction = worker_stat1.get('personal_satisfaction_score', 50)
        
        for worker_id2, worker_stat2 in stats['worker_stats'].items():
            if worker_id1 == worker_id2:
                envy_matrix[worker_id1][worker_id2] = 0  # No self-envy
                continue
            
            # Calculate if worker1 would prefer worker2's assignment
            worker2_satisfaction = worker_stat2.get('personal_satisfaction_score', 50)
            worker2_hours = worker_stat2['total_hours']
            worker1_hours = worker_stat1['total_hours']
            
            # Envy calculation: prefer less work OR higher satisfaction with similar work
            envy_score = 0
            if worker2_hours < worker1_hours and worker2_satisfaction >= worker1_satisfaction:
                envy_score = (worker1_hours - worker2_hours) * 10  # Envy for less work
            elif worker2_satisfaction > worker1_satisfaction + 10 and abs(worker2_hours - worker1_hours) < 5:
                envy_score = worker2_satisfaction - worker1_satisfaction  # Envy for better satisfaction
            
            envy_matrix[worker_id1][worker_id2] = max(0, envy_score)

    # 4. EGALITARIAN METRICS (Minimize maximum burden)
    if stats['worker_stats']:
        max_burden = max(w['weighted_difficulty_burden'] for w in stats['worker_stats'].values())
        min_burden = min(w['weighted_difficulty_burden'] for w in stats['worker_stats'].values())
        burden_variance = statistics.variance([w['weighted_difficulty_burden'] for w in stats['worker_stats'].values()])
        
        stats['fairness_metrics']['max_burden'] = max_burden
        stats['fairness_metrics']['min_burden'] = min_burden
        stats['fairness_metrics']['burden_variance'] = burden_variance
        stats['fairness_metrics']['egalitarian_score'] = max(0, 100 - (burden_variance * 10))  # Lower variance = higher score

    # 5. NIGHT SHIFT FAIRNESS
    night_shifts_total = sum(w['night_shifts'] for w in stats['worker_stats'].values())
    active_workers = len([w for w in stats['worker_stats'].values() if w['total_shifts'] > 0])
    
    if active_workers > 0 and night_shifts_total > 0:
        fair_night_share = night_shifts_total / active_workers
        night_fairness_scores = {}
        
        for worker_id, worker_stat in stats['worker_stats'].items():
            if worker_stat['total_shifts'] > 0:
                night_deviation = abs(worker_stat['night_shifts'] - fair_night_share)
                night_fairness_scores[worker_id] = max(0, 100 - (night_deviation * 20))
            else:
                night_fairness_scores[worker_id] = 100
        
        stats['fairness_metrics']['night_fairness_scores'] = night_fairness_scores
        stats['fairness_metrics']['average_night_fairness'] = statistics.mean(night_fairness_scores.values())

    # 6. REST TIME ANALYSIS
    for worker_id, worker_stat in stats['worker_stats'].items():
        assignments = sorted(worker_stat['assignments'], key=lambda x: x['start'])
        rest_periods = []
        max_continuous = 0
        current_continuous = 0
        
        for i in range(len(assignments) - 1):
            current_end = assignments[i]['end']
            next_start = assignments[i + 1]['start']
            rest_duration = (next_start - current_end).total_seconds() / 3600.0
            
            if rest_duration > 0:
                rest_periods.append(rest_duration)
                current_continuous = 0
            else:
                current_continuous += assignments[i]['duration']
                max_continuous = max(max_continuous, current_continuous)
        
        worker_stat['rest_periods'] = rest_periods
        worker_stat['max_continuous_work'] = max_continuous
        worker_stat['average_rest_time'] = statistics.mean(rest_periods) if rest_periods else 0
        worker_stat['min_rest_time'] = min(rest_periods) if rest_periods else 0

    # Store enhanced metrics
    stats['satisfaction_metrics'] = {
        'avg_satisfaction': statistics.mean([w.get('personal_satisfaction_score', 50) for w in stats['worker_stats'].values()]),
        'satisfaction_variance': statistics.variance([w.get('personal_satisfaction_score', 50) for w in stats['worker_stats'].values()]),
        'proportional_scores': {w['name']: w.get('proportional_fairness_score', 0) for w in stats['worker_stats'].values()}
    }
    
    stats['envy_metrics'] = {
        'envy_matrix': {stats['worker_stats'][w1]['name']: {stats['worker_stats'][w2]['name']: score 
                       for w2, score in envy_scores.items()} 
                       for w1, envy_scores in envy_matrix.items()},
        'total_envy': sum(sum(scores.values()) for scores in envy_matrix.values()),
        'envy_free': sum(sum(scores.values()) for scores in envy_matrix.values()) == 0
    }

    # --- Generate Enhanced Insights ---
    insights = []
    
    # Satisfaction insights
    satisfaction_scores = [w.get('personal_satisfaction_score', 50) for w in stats['worker_stats'].values()]
    avg_satisfaction = statistics.mean(satisfaction_scores)
    
    if avg_satisfaction < 40:
        insights.append("⚠️ <b>Low overall worker satisfaction</b> - workers are getting more difficult tasks than expected")
    elif avg_satisfaction > 80:
        insights.append("✅ <b>High worker satisfaction</b> - workers are getting appropriately matched tasks")
    
    # Envy insights
    if stats['envy_metrics']['envy_free']:
        insights.append("🎯 <b>Envy-free allocation achieved</b> - no worker prefers another's assignment")
    else:
        total_envy = stats['envy_metrics']['total_envy']
        if total_envy > 50:
            insights.append(f"⚠️ <b>High envy detected</b> - total envy score: {total_envy:.1f}")
    
    # Egalitarian insights
    if 'egalitarian_score' in stats['fairness_metrics']:
        egal_score = stats['fairness_metrics']['egalitarian_score']
        if egal_score > 80:
            insights.append("⚖️ <b>Good egalitarian distribution</b> - workload burdens are well balanced")
        elif egal_score < 50:
            insights.append("⚠️ <b>Uneven burden distribution</b> - some workers have significantly harder workloads")
    
    # Night shift insights
    if 'average_night_fairness' in stats['fairness_metrics']:
        night_fairness = stats['fairness_metrics']['average_night_fairness']
        if night_fairness > 80:
            insights.append("🌙 <b>Fair night shift distribution</b>")
        elif night_fairness < 60:
            insights.append("⚠️ <b>Uneven night shift distribution</b> - some workers have disproportionate night work")
    
    # Rest time insights
    min_rest_times = [w['min_rest_time'] for w in stats['worker_stats'].values() if w['rest_periods']]
    if min_rest_times and min(min_rest_times) < 8:
        insights.append("😴 <b>Insufficient rest periods detected</b> - some workers have less than 8 hours between shifts")
    
    # Proportional fairness insights
    prop_scores = [w.get('proportional_fairness_score', 0) for w in stats['worker_stats'].values()]
    avg_prop = statistics.mean(prop_scores) if prop_scores else 0
    if avg_prop < 70:
        insights.append("📊 <b>Disproportional allocation</b> - work distribution doesn't match worker qualifications")

    return render_template('fairness_statistics.html', period=period, stats=stats, insights=insights)





############### Adding CSV and Excel difficulty export for workers ###############
# Add these routes to routes.py

# Add these updated functions to your routes.py file

@main_bp.route('/period/<int:period_id>/export_rating_template')
def export_rating_template(period_id):
    """Export CSV template in matrix format: Workers as rows, Job Roles as columns"""
    period = SchedulingPeriod.query.get_or_404(period_id)
    
    # Get all workers and job roles for this period
    all_workers = Worker.query.options(selectinload(Worker.qualified_roles)).all()
    job_roles = JobRole.query.filter_by(scheduling_period_id=period.id).order_by(JobRole.name).all()
    
    if not all_workers:
        flash("No workers found. Add workers before exporting rating template.", "warning")
        return redirect(url_for('main.manage_job_roles_for_period', period_id=period.id))
    
    if not job_roles:
        flash("No job roles found for this period.", "warning")
        return redirect(url_for('main.manage_job_roles_for_period', period_id=period.id))
    
    # Create CSV data in matrix format
    csv_rows = []
    
    # Create header row: Worker Name + all job role names
    header_row = ['Worker Name'] + [role.name for role in job_roles]
    csv_rows.append(header_row)
    
    # Add instruction row
    instruction_row = ['INSTRUCTIONS: Rate 1-5 (1=Easy, 5=Very Hard). Leave empty if not qualified or don\'t want to rate.'] + [''] * len(job_roles)
    csv_rows.append(instruction_row)
    
    # Add current difficulty reference row
    current_diff_row = ['Current System Difficulty'] + [f"{role.difficulty_multiplier:.1f}" for role in job_roles]
    csv_rows.append(current_diff_row)
    
    # Add qualification reference row 
    qualified_workers_per_role = []
    for role in job_roles:
        qualified_count = sum(1 for worker in all_workers if role in worker.qualified_roles)
        qualified_workers_per_role.append(f"{qualified_count} qualified")
    
    qualification_row = ['Workers Qualified'] + qualified_workers_per_role
    csv_rows.append(qualification_row)
    
    # Add empty separator row
    csv_rows.append([''] * len(header_row))
    
    # Add data rows: one row per worker
    for worker in all_workers:
        worker_row = [worker.name]
        
        for role in job_roles:
            is_qualified = role in worker.qualified_roles
            # Pre-fill with empty string - workers will fill in ratings
            # Add a hint in the cell if not qualified
            if is_qualified:
                worker_row.append('')  # Empty for qualified workers to fill
            else:
                worker_row.append('N/A')  # Mark as N/A for unqualified workers
        
        csv_rows.append(worker_row)
    
    # Create CSV content
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerows(csv_rows)
    
    # Create response
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = f'attachment; filename=difficulty_matrix_{period.name.replace(" ", "_")}_{datetime.now().strftime("%Y%m%d")}.csv'
    
    flash(f"Matrix format rating template exported with {len(all_workers)} workers and {len(job_roles)} job roles. Each worker gets one row to fill.", "success")
    return response


# @main_bp.route('/period/<int:period_id>/import_ratings', methods=['GET', 'POST'])
# def import_worker_ratings(period_id):
#     """Import worker difficulty ratings from matrix format CSV"""
#     period = SchedulingPeriod.query.get_or_404(period_id)
    
#     if request.method == 'GET':
#         # Show import form
#         return render_template('import_ratings.html', period=period)
    
#     # Handle POST request with file upload
#     if 'rating_file' not in request.files:
#         flash('No file selected for upload.', 'danger')
#         return redirect(url_for('main.import_worker_ratings', period_id=period_id))
    
#     file = request.files['rating_file']
#     if file.filename == '':
#         flash('No file selected for upload.', 'danger')
#         return redirect(url_for('main.import_worker_ratings', period_id=period_id))
    
#     if not file.filename.lower().endswith('.csv'):
#         flash('Please upload a CSV file.', 'danger')
#         return redirect(url_for('main.import_worker_ratings', period_id=period_id))
    
#     try:
#         # Read CSV content
#         csv_content = file.read().decode('utf-8')
#         csv_lines = csv_content.strip().split('\n')
        
#         if len(csv_lines) < 6:  # Header + instruction + current diff + qualified + empty + at least 1 worker
#             flash('CSV file appears to be empty or invalid. Please use the exported matrix template.', 'danger')
#             return redirect(url_for('main.import_worker_ratings', period_id=period_id))
        
#         # Parse CSV
#         reader = csv.reader(io.StringIO(csv_content))
#         rows = list(reader)
        
#         # Extract header row (job role names)
#         header_row = rows[0]
#         if len(header_row) < 2 or header_row[0] != 'Worker Name':
#             flash('Invalid CSV format. Please use the exported matrix template.', 'danger')
#             return redirect(url_for('main.import_worker_ratings', period_id=period_id))
        
#         job_role_names = header_row[1:]  # Skip "Worker Name" column
        
#         # Find where the actual data starts (skip instruction rows)
#         data_start_row = 5  # Skip header + instruction + current diff + qualified + empty
#         worker_rows = rows[data_start_row:]
        
#         # Process ratings
#         ratings_by_role = defaultdict(list)  # role_name -> list of ratings
#         processed_count = 0
#         skipped_count = 0
#         error_count = 0
        
#         for row_num, row in enumerate(worker_rows, start=data_start_row + 1):
#             if len(row) < 2:  # Skip empty rows
#                 continue
                
#             worker_name = row[0].strip()
#             if not worker_name or worker_name.startswith('INSTRUCTIONS'):
#                 continue
            
#             # Process each role rating for this worker
#             for col_idx, role_name in enumerate(job_role_names):
#                 if col_idx + 1 >= len(row):  # Skip if not enough columns
#                     continue
                    
#                 rating_str = row[col_idx + 1].strip()
                
#                 # Skip empty, N/A, or non-numeric ratings
#                 if not rating_str or rating_str.upper() in ['N/A', 'NA', '']:
#                     skipped_count += 1
#                     continue
                
#                 try:
#                     rating = float(rating_str)
                    
#                     # Normalize and convert to integer
#                     if rating < 1:
#                         rating = 1
#                     elif rating > 5:
#                         rating = 5
                    
#                     rating = int(round(rating))  # Convert to integer
                    
#                     ratings_by_role[role_name].append({
#                         'worker': worker_name,
#                         'rating': rating,
#                         'comments': f"Matrix import from {worker_name}"
#                     })
#                     processed_count += 1
                    
#                 except ValueError:
#                     current_app.logger.warning(f"Row {row_num}: Invalid rating '{rating_str}' for {worker_name}-{role_name}")
#                     error_count += 1
#                     continue
        
#         if not ratings_by_role:
#             flash('No valid ratings found in the uploaded matrix. Please check the format and try again.', 'warning')
#             return redirect(url_for('main.import_worker_ratings', period_id=period_id))
        
#         # Calculate average ratings and update job roles
#         updated_roles = []
#         role_stats = {}
        
#         for role_name, ratings_list in ratings_by_role.items():
#             # Find the corresponding job role
#             role = JobRole.query.filter_by(
#                 scheduling_period_id=period_id, 
#                 name=role_name
#             ).first()
            
#             if not role:
#                 current_app.logger.warning(f"Role '{role_name}' not found in period {period.name}")
#                 continue
            
#             # Calculate average rating
#             rating_values = [r['rating'] for r in ratings_list]
#             avg_rating = sum(rating_values) / len(rating_values)
#             old_difficulty = role.difficulty_multiplier
            
#             # Update role difficulty (round to 2 decimal places)
#             role.difficulty_multiplier = round(avg_rating, 2)
#             updated_roles.append(role)
            
#             role_stats[role_name] = {
#                 'old_difficulty': old_difficulty,
#                 'new_difficulty': role.difficulty_multiplier,
#                 'num_ratings': len(rating_values),
#                 'ratings': rating_values,
#                 'avg_rating': avg_rating
#             }
        
#         # Commit changes
#         db.session.commit()
        
#         # Create success message with details
#         flash(f"Successfully imported matrix ratings! Updated {len(updated_roles)} job roles based on {processed_count} worker ratings.", "success")
        
#         if skipped_count > 0:
#             flash(f"Skipped {skipped_count} empty/N/A ratings (workers who didn't rate certain roles).", "info")
        
#         if error_count > 0:
#             flash(f"Found {error_count} invalid ratings that were ignored.", "warning")
        
#         # Store detailed results in session for display
#         session['import_results'] = {
#             'role_stats': role_stats,
#             'processed_count': processed_count,
#             'skipped_count': skipped_count,
#             'error_count': error_count
#         }
        
#         # return redirect(url_for('main.show_import_results', period_id=period_id))
#         return redirect(url_for('main.manage_job_roles_for_period', period_id=period_id))
        
#     except Exception as e:
#         current_app.logger.error(f"Error importing matrix ratings: {e}")
#         flash(f"Error processing matrix file: {e}", "danger")
#         return redirect(url_for('main.import_worker_ratings', period_id=period_id))






############## Integrating extreme rating pattern detection ##############
@main_bp.route('/period/<int:period_id>/import_ratings', methods=['GET', 'POST'])
def import_worker_ratings(period_id):
    """Import worker difficulty ratings from matrix format CSV with extreme rating protection"""
    period = SchedulingPeriod.query.get_or_404(period_id)
    
    if request.method == 'GET':
        # Show import form
        return render_template('import_ratings.html', period=period)
    
    # Handle POST request with file upload
    if 'rating_file' not in request.files:
        flash('No file selected for upload.', 'danger')
        return redirect(url_for('main.import_worker_ratings', period_id=period_id))
    
    file = request.files['rating_file']
    if file.filename == '':
        flash('No file selected for upload.', 'danger')
        return redirect(url_for('main.import_worker_ratings', period_id=period_id))
    
    if not file.filename.lower().endswith('.csv'):
        flash('Please upload a CSV file.', 'danger')
        return redirect(url_for('main.import_worker_ratings', period_id=period_id))
    
    try:
        # Read CSV content
        csv_content = file.read().decode('utf-8')
        csv_lines = csv_content.strip().split('\n')
        
        if len(csv_lines) < 6:  # Header + instruction + current diff + qualified + empty + at least 1 worker
            flash('CSV file appears to be empty or invalid. Please use the exported matrix template.', 'danger')
            return redirect(url_for('main.import_worker_ratings', period_id=period_id))
        
        # Parse CSV
        reader = csv.reader(io.StringIO(csv_content))
        rows = list(reader)
        
        # Extract header row (job role names)
        header_row = rows[0]
        if len(header_row) < 2 or header_row[0] != 'Worker Name':
            flash('Invalid CSV format. Please use the exported matrix template.', 'danger')
            return redirect(url_for('main.import_worker_ratings', period_id=period_id))
        
        job_role_names = header_row[1:]  # Skip "Worker Name" column
        
        # Find where the actual data starts (skip instruction rows)
        data_start_row = 5  # Skip header + instruction + current diff + qualified + empty
        worker_rows = rows[data_start_row:]
        
        # Process ratings
        ratings_by_role = defaultdict(list)  # role_name -> list of ratings
        processed_count = 0
        skipped_count = 0
        error_count = 0
        
        for row_num, row in enumerate(worker_rows, start=data_start_row + 1):
            if len(row) < 2:  # Skip empty rows
                continue
                
            worker_name = row[0].strip()
            if not worker_name or worker_name.startswith('INSTRUCTIONS'):
                continue
            
            # Process each role rating for this worker
            for col_idx, role_name in enumerate(job_role_names):
                if col_idx + 1 >= len(row):  # Skip if not enough columns
                    continue
                    
                rating_str = row[col_idx + 1].strip()
                
                # Skip empty, N/A, or non-numeric ratings
                if not rating_str or rating_str.upper() in ['N/A', 'NA', '']:
                    skipped_count += 1
                    continue
                
                try:
                    rating = float(rating_str)
                    
                    # Normalize and convert to integer
                    if rating < 1:
                        rating = 1
                    elif rating > 5:
                        rating = 5
                    
                    rating = int(round(rating))  # Convert to integer
                    
                    ratings_by_role[role_name].append({
                        'worker': worker_name,
                        'rating': rating,
                        'comments': f"Matrix import from {worker_name}"
                    })
                    processed_count += 1
                    
                except ValueError:
                    current_app.logger.warning(f"Row {row_num}: Invalid rating '{rating_str}' for {worker_name}-{role_name}")
                    error_count += 1
                    continue
        
        if not ratings_by_role:
            flash('No valid ratings found in the uploaded matrix. Please check the format and try again.', 'warning')
            return redirect(url_for('main.import_worker_ratings', period_id=period_id))
        
        # ENHANCED: Detect and handle extreme rating patterns
        pattern_warnings, cleaned_ratings, normalized_ratings = detect_extreme_rating_patterns(ratings_by_role, period_id)
        distribution_warnings = analyze_role_distribution(cleaned_ratings if cleaned_ratings else ratings_by_role)
        
        # Use Strategy 1: Remove extreme workers entirely (recommended approach)
        final_ratings = cleaned_ratings if cleaned_ratings else ratings_by_role
        
        # Combine all warnings
        all_warnings = pattern_warnings + distribution_warnings
        
        # Calculate average ratings and update job roles
        updated_roles = []
        role_stats = {}
        removed_roles = []
        
        for role_name, ratings_list in final_ratings.items():
            # Find the corresponding job role
            role = JobRole.query.filter_by(
                scheduling_period_id=period_id, 
                name=role_name
            ).first()
            
            if not role:
                current_app.logger.warning(f"Role '{role_name}' not found in period {period.name}")
                continue
            
            if not ratings_list:  # No valid ratings after cleaning
                current_app.logger.warning(f"No valid ratings for role '{role_name}' after cleaning")
                removed_roles.append(role_name)
                continue
            
            # Calculate average rating
            rating_values = [r['rating'] for r in ratings_list]
            avg_rating = sum(rating_values) / len(rating_values)
            old_difficulty = role.difficulty_multiplier
            
            # Update role difficulty (round to 2 decimal places)
            role.difficulty_multiplier = round(avg_rating, 2)
            updated_roles.append(role)
            
            role_stats[role_name] = {
                'old_difficulty': old_difficulty,
                'new_difficulty': role.difficulty_multiplier,
                'num_ratings': len(rating_values),
                'ratings': rating_values,
                'avg_rating': avg_rating,
                'cleaned': len(ratings_list) != len(ratings_by_role.get(role_name, []))
            }
        
        # Handle roles that lost all ratings due to cleaning
        for role_name in removed_roles:
            if role_name in ratings_by_role:  # Had ratings before cleaning
                all_warnings.append(f"Role '{role_name}' lost all ratings after removing extreme patterns - keeping original difficulty")
        
        # Commit changes
        db.session.commit()
        
        # Create success message with details
        flash(f"Successfully imported matrix ratings! Updated {len(updated_roles)} job roles based on {processed_count} worker ratings.", "success")
        
        if skipped_count > 0:
            flash(f"Skipped {skipped_count} empty/N/A ratings (workers who didn't rate certain roles).", "info")
        
        if error_count > 0:
            flash(f"Found {error_count} invalid ratings that were ignored.", "warning")
        
        # Show pattern detection warnings
        for warning in all_warnings:
            flash(warning, "warning")
        
        # Store detailed results in session for display
        session['import_results'] = {
            'role_stats': role_stats,
            'processed_count': processed_count,
            'skipped_count': skipped_count,
            'error_count': error_count,
            'pattern_warnings': pattern_warnings,
            'distribution_warnings': distribution_warnings,
            'extreme_workers_detected': len(pattern_warnings) > 0
        }
        
        return redirect(url_for('main.show_import_results', period_id=period_id))
        
    except Exception as e:
        current_app.logger.error(f"Error importing matrix ratings: {e}")
        flash(f"Error processing matrix file: {e}", "danger")
        return redirect(url_for('main.import_worker_ratings', period_id=period_id))


# Helper functions - add these to your routes.py file as well:

def detect_extreme_rating_patterns(ratings_by_role, period_id):
    """
    Detect and handle extreme rating patterns that could game the system
    Returns: warnings, normalized_ratings_by_role
    """
    warnings = []
    normalized_ratings = {}
    
    # Analyze individual worker rating patterns
    worker_ratings = defaultdict(list)  # worker -> [all their ratings]
    
    # Collect all ratings by worker
    for role_name, ratings_list in ratings_by_role.items():
        for rating_data in ratings_list:
            worker_ratings[rating_data['worker']].append(rating_data['rating'])
    
    # Detect extreme patterns
    extreme_workers = []
    for worker_name, worker_rating_list in worker_ratings.items():
        if len(worker_rating_list) < 2:  # Need at least 2 ratings to analyze
            continue
            
        rating_variance = statistics.variance(worker_rating_list) if len(worker_rating_list) > 1 else 0
        rating_mean = statistics.mean(worker_rating_list)
        
        # Pattern 1: All ratings the same (variance = 0)
        if rating_variance == 0:
            if rating_mean <= 1.5:
                warnings.append(f"{worker_name} rated ALL roles as {int(rating_mean)} (too easy) - ratings removed to prevent overwork")
                extreme_workers.append((worker_name, 'all_easy'))
            elif rating_mean >= 4.5:
                warnings.append(f"{worker_name} rated ALL roles as {int(rating_mean)} (too hard) - ratings removed to prevent gaming")
                extreme_workers.append((worker_name, 'all_hard'))
        
        # Pattern 2: Very low variance (almost all the same)
        elif rating_variance < 0.3 and len(worker_rating_list) >= 3:
            warnings.append(f"{worker_name} gave very similar ratings (variance: {rating_variance:.2f}) - may indicate unrealistic assessment")
            extreme_workers.append((worker_name, 'low_variance'))
    
    # Remove extreme ratings entirely
    cleaned_ratings = {}
    for role_name, ratings_list in ratings_by_role.items():
        cleaned_list = []
        for rating_data in ratings_list:
            worker_name = rating_data['worker']
            # Skip ratings from workers with extreme patterns
            if not any(worker_name == extreme[0] for extreme in extreme_workers):
                cleaned_list.append(rating_data)
        
        if cleaned_list:  # Only include roles that have valid ratings left
            cleaned_ratings[role_name] = cleaned_list
    
    return warnings, cleaned_ratings, normalized_ratings


def analyze_role_distribution(ratings_by_role):
    """Check if roles have meaningful difficulty differences"""
    if len(ratings_by_role) < 2:
        return []
    
    warnings = []
    role_averages = {}
    
    # Calculate average for each role
    for role_name, ratings_list in ratings_by_role.items():
        if ratings_list:
            avg = sum(r['rating'] for r in ratings_list) / len(ratings_list)
            role_averages[role_name] = avg
    
    if len(role_averages) < 2:
        return warnings
    
    # Check if all roles have similar difficulty
    avg_values = list(role_averages.values())
    difficulty_variance = statistics.variance(avg_values)
    
    if difficulty_variance < 0.5:  # Very low variance between roles
        warnings.append(f"All roles have similar difficulty ratings (variance: {difficulty_variance:.2f}) - difficulty weighting may not be very effective")
    
    # List roles by difficulty for reference
    sorted_roles = sorted(role_averages.items(), key=lambda x: x[1])
    if len(sorted_roles) >= 3:
        easiest = sorted_roles[0]
        hardest = sorted_roles[-1]
        warnings.append(f"Easiest role: {easiest[0]} ({easiest[1]:.1f}), Hardest role: {hardest[0]} ({hardest[1]:.1f})")
    
    return warnings




@main_bp.route('/period/<int:period_id>/import_results')
def show_import_results(period_id):
    """Show detailed results of rating import"""
    period = SchedulingPeriod.query.get_or_404(period_id)
    
    import_results = session.pop('import_results', None)
    if not import_results:
        flash("No import results to display.", "warning")
        return redirect(url_for('main.manage_job_roles_for_period', period_id=period_id))
    
    return render_template('import_results.html', 
                         period=period, 
                         results=import_results)


@main_bp.route('/period/<int:period_id>/reset_difficulties', methods=['POST'])
def reset_difficulties(period_id):
    """Reset all job role difficulties to 1.0 (neutral)"""
    period = SchedulingPeriod.query.get_or_404(period_id)
    
    job_roles = JobRole.query.filter_by(scheduling_period_id=period_id).all()
    if not job_roles:
        flash("No job roles found to reset.", "warning")
        return redirect(url_for('main.manage_job_roles_for_period', period_id=period_id))
    
    # Reset all difficulties to 1.0
    for role in job_roles:
        role.difficulty_multiplier = 1.0
    
    db.session.commit()
    
    flash(f"Reset difficulty ratings for {len(job_roles)} job roles to neutral (1.0).", "success")
    return redirect(url_for('main.manage_job_roles_for_period', period_id=period_id))












##################### Handling rating scenarios ###########################

import statistics


def detect_extreme_rating_patterns(ratings_by_role, period_id):
    """
    Detect and handle extreme rating patterns that could game the system
    Returns: warnings, normalized_ratings_by_role
    """
    warnings = []
    normalized_ratings = {}
    
    # Analyze individual worker rating patterns
    worker_ratings = defaultdict(list)  # worker -> [all their ratings]
    
    # Collect all ratings by worker
    for role_name, ratings_list in ratings_by_role.items():
        for rating_data in ratings_list:
            worker_ratings[rating_data['worker']].append(rating_data['rating'])
    
    # Detect extreme patterns
    extreme_workers = []
    for worker_name, worker_rating_list in worker_ratings.items():
        if len(worker_rating_list) < 2:  # Need at least 2 ratings to analyze
            continue
            
        rating_variance = statistics.variance(worker_rating_list) if len(worker_rating_list) > 1 else 0
        rating_mean = statistics.mean(worker_rating_list)
        
        # Pattern 1: All ratings the same (variance = 0)
        if rating_variance == 0:
            if rating_mean <= 1.5:
                warnings.append(f"⚠️ {worker_name} rated ALL roles as {int(rating_mean)} (too easy) - this may cause overwork")
                extreme_workers.append((worker_name, 'all_easy'))
            elif rating_mean >= 4.5:
                warnings.append(f"⚠️ {worker_name} rated ALL roles as {int(rating_mean)} (too hard) - this may avoid work")
                extreme_workers.append((worker_name, 'all_hard'))
        
        # Pattern 2: Very low variance (almost all the same)
        elif rating_variance < 0.3:
            warnings.append(f"⚠️ {worker_name} gave very similar ratings (variance: {rating_variance:.2f}) - may indicate gaming")
            extreme_workers.append((worker_name, 'low_variance'))
    
    # Option 1: Remove extreme ratings entirely
    cleaned_ratings = {}
    for role_name, ratings_list in ratings_by_role.items():
        cleaned_list = []
        for rating_data in ratings_list:
            worker_name = rating_data['worker']
            # Skip ratings from workers with extreme patterns
            if not any(worker_name == extreme[0] for extreme in extreme_workers):
                cleaned_list.append(rating_data)
        
        if cleaned_list:  # Only include roles that have valid ratings left
            cleaned_ratings[role_name] = cleaned_list
        else:
            # If no valid ratings left, warn and use default
            warnings.append(f"⚠️ Role '{role_name}' has no valid ratings after removing extreme patterns - using default 1.0")
    
    # Option 2: Normalize ratings within each worker to force distribution
    normalized_ratings = apply_worker_normalization(ratings_by_role, extreme_workers)
    
    return warnings, cleaned_ratings, normalized_ratings


def apply_worker_normalization(ratings_by_role, extreme_workers):
    """
    Force each worker's ratings to have proper distribution (not all the same)
    """
    # Collect all ratings by worker first
    worker_all_ratings = defaultdict(dict)  # worker -> {role: rating}
    
    for role_name, ratings_list in ratings_by_role.items():
        for rating_data in ratings_list:
            worker_name = rating_data['worker']
            worker_all_ratings[worker_name][role_name] = rating_data['rating']
    
    # Normalize extreme workers
    normalized_ratings_by_role = defaultdict(list)
    
    for role_name, ratings_list in ratings_by_role.items():
        for rating_data in ratings_list:
            worker_name = rating_data['worker']
            original_rating = rating_data['rating']
            
            # Check if this worker has extreme patterns
            is_extreme = any(worker_name == extreme[0] for extreme in extreme_workers)
            
            if is_extreme:
                # Apply forced distribution: convert flat ratings to relative rankings
                worker_ratings = worker_all_ratings[worker_name]
                
                if len(worker_ratings) >= 3:  # Need at least 3 roles to rank
                    # Convert to percentile-based ratings
                    sorted_roles = sorted(worker_ratings.items(), key=lambda x: x[1])
                    percentile_rating = get_percentile_rating(role_name, sorted_roles)
                    normalized_rating = percentile_rating
                else:
                    # Not enough roles to normalize, use default
                    normalized_rating = 3.0  # Default to moderate
                
                normalized_ratings_by_role[role_name].append({
                    'worker': worker_name,
                    'rating': normalized_rating,
                    'comments': f"Normalized from {original_rating} (extreme pattern detected)"
                })
            else:
                # Keep original rating for normal workers
                normalized_ratings_by_role[role_name].append(rating_data)
    
    return dict(normalized_ratings_by_role)


def get_percentile_rating(target_role, sorted_roles):
    """Convert role position in worker's ranking to 1-5 scale"""
    total_roles = len(sorted_roles)
    
    # Find position of target role
    position = next(i for i, (role, _) in enumerate(sorted_roles) if role == target_role)
    
    # Convert position to 1-5 scale
    percentile = position / (total_roles - 1) if total_roles > 1 else 0.5
    
    # Map percentile to rating scale
    if percentile <= 0.2:
        return 1  # Bottom 20% = Easy
    elif percentile <= 0.4:
        return 2  # Next 20% = Light  
    elif percentile <= 0.6:
        return 3  # Middle 20% = Moderate
    elif percentile <= 0.8:
        return 4  # Next 20% = Hard
    else:
        return 5  # Top 20% = Very Hard


def analyze_role_distribution(ratings_by_role):
    """Check if roles have meaningful difficulty differences"""
    if len(ratings_by_role) < 2:
        return []
    
    warnings = []
    role_averages = {}
    
    # Calculate average for each role
    for role_name, ratings_list in ratings_by_role.items():
        if ratings_list:
            avg = sum(r['rating'] for r in ratings_list) / len(ratings_list)
            role_averages[role_name] = avg
    
    if len(role_averages) < 2:
        return warnings
    
    # Check if all roles have similar difficulty
    avg_values = list(role_averages.values())
    difficulty_variance = statistics.variance(avg_values)
    
    if difficulty_variance < 0.5:  # Very low variance between roles
        warnings.append(f"⚠️ All roles have similar difficulty ratings (variance: {difficulty_variance:.2f}) - difficulty weighting may not be effective")
    
    # List roles by difficulty for reference
    sorted_roles = sorted(role_averages.items(), key=lambda x: x[1])
    if len(sorted_roles) >= 3:
        easiest = sorted_roles[0]
        hardest = sorted_roles[-1]
        warnings.append(f"ℹ️ Easiest role: {easiest[0]} ({easiest[1]:.1f}), Hardest role: {hardest[0]} ({hardest[1]:.1f})")
    
    return warnings


# Update your import_worker_ratings function to use these validations:

# Replace the existing processing section with this enhanced version:
def enhanced_rating_processing(ratings_by_role, period_id):
    """Enhanced processing with extreme pattern detection"""
    
    # Detect extreme patterns
    pattern_warnings, cleaned_ratings, normalized_ratings = detect_extreme_rating_patterns(ratings_by_role, period_id)
    
    # Analyze role distribution
    distribution_warnings = analyze_role_distribution(cleaned_ratings)
    
    # Combine all warnings
    all_warnings = pattern_warnings + distribution_warnings
    
    # Use cleaned ratings (with extreme patterns removed)
    final_ratings = cleaned_ratings
    
    # Alternative: Use normalized ratings (extreme patterns fixed)
    # final_ratings = normalized_ratings
    
    return final_ratings, all_warnings